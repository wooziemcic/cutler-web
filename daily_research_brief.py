from __future__ import annotations

import json
import re
import shutil
import tempfile
import zipfile
from collections import Counter
from datetime import datetime
from pathlib import Path, PurePosixPath
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd


CATEGORIES = ("Banks", "Companies", "Credit", "Green Street", "Research")
SUPPORTED_TEXT_EXTENSIONS = {".pdf", ".xlsx", ".xls", ".txt", ".md", ".csv"}
JUNK_NAMES = {".ds_store", "thumbs.db", "desktop.ini"}
JUNK_PARTS = {"__macosx", ".git", ".svn", "__pycache__"}

BROKER_PATTERNS = [
    ("Morgan Stanley", r"\bMORGAN\s+STANLEY\b"),
    ("Gimme Credit", r"\bGIMME\s+CREDIT\b"),
    ("Green Street", r"\bGREEN\s+STREET\b"),
    ("Morningstar", r"\bMORNINGSTAR\b"),
    ("JPMorgan", r"\b(?:JPM|J\.?\s*P\.?\s*MORGAN|JP\s*MORGAN|JPMORGAN)\b"),
    ("Goldman Sachs", r"\b(?:GS|GOLDMAN(?:\s+SACHS)?)\b"),
    ("Wells Fargo", r"\b(?:WF|WELLS\s+FARGO)\b"),
    ("Bank of America", r"\b(?:BOFA|BOF\s+A|BANK\s+OF\s+AMERICA|BAML)\b"),
    ("Barclays", r"\bBARCLAYS\b"),
    ("RBC", r"\bRBC\b"),
    ("Jefferies", r"\bJEFFERIES\b"),
    ("Hovde", r"\bHOVDE\b"),
    ("Brean", r"\bBREAN\b"),
    ("Evercore", r"\bEVERCORE\b"),
    ("Mizuho", r"\bMIZUHO\b"),
    ("Guggenheim", r"\bGUGGENHEIM\b"),
    ("Stifel", r"\bSTIFEL\b"),
    ("Deutsche", r"\bDEUTSCHE\b"),
    ("Cantor", r"\bCANTOR\b"),
    ("Citizens", r"\bCITIZENS\b"),
    ("BMO", r"\bBMO\b"),
    ("UBS", r"\bUBS\b"),
    ("Citi", r"\b(?:CITI|CITIGROUP)\b"),
    ("Truist", r"\bTRUIST\b"),
    ("Piper", r"\bPIPER\b"),
    ("Wedbush", r"\bWEDBUSH\b"),
    ("Needham", r"\bNEEDHAM\b"),
    ("Oppenheimer", r"\bOPPENHEIMER\b"),
    ("Raymond James", r"\bRAYMOND\s+JAMES\b"),
    ("TD Cowen", r"\b(?:TD\s+COWEN|COWEN)\b"),
    ("KeyBanc", r"\bKEY\s*BANC\b"),
    ("Wolfe", r"\bWOLFE\b"),
    ("Baird", r"\bBAIRD\b"),
    ("Stephens", r"\bSTEPHENS\b"),
]

TOP_TIER_BROKERS = {
    "JPMorgan", "Goldman Sachs", "Wells Fargo", "Bank of America", "Morgan Stanley",
    "Barclays", "Citi", "UBS", "Deutsche", "RBC",
}
SPECIALTY_SOURCES = {
    "Gimme Credit", "Green Street", "Morningstar", "Hovde", "Brean", "Needham",
    "Oppenheimer", "Raymond James", "TD Cowen", "KeyBanc", "Wolfe", "Baird", "Stephens",
}

DOCUMENT_TYPES = [
    ("earnings presentation", (r"\bearnings?\s+presentation\b", r"\binvestor\s+presentation\b")),
    ("earnings supplement", (r"\bearnings?\s+supplement\b",)),
    ("earnings release", (r"\bearnings?\s+release\b", r"\bresults?\s+release\b")),
    ("prepared remarks", (r"\bprepared\s+remarks?\b",)),
    ("transcript", (r"\btranscript\b", r"\bearnings?\s+call\b")),
    ("10-Q/10Q", (r"\b10[\s_-]?q\b",)),
    ("8-K/8K", (r"\b8[\s_-]?k\b",)),
    ("current report", (r"\bcurrent\s+report\b",)),
    ("credit report", (r"\bcredit\s+(?:report|update|research)\b",)),
    ("analyst report", (r"\banalyst\s+report\b", r"\bequity\s+research\b")),
    ("sector report", (r"\bsector\s+(?:report|update|outlook)\b",)),
    ("industry report", (r"\bindustry\s+(?:report|update|outlook)\b",)),
    ("company overview", (r"\bcompany\s+overview\b", r"\bcompany\s+profile\b")),
    ("data supplement", (r"\bdata\s+supplement\b",)),
    ("MD&A", (r"\bmd\s*&\s*a\b", r"\bmanagement.?s discussion\b")),
    ("press release", (r"\bpress\s+release\b",)),
    ("dividend announcement", (r"\bdividend\b",)),
    ("insider trading", (r"\binsider\s+(?:trading|transaction|buy|sale)\b",)),
    ("fear and greed", (r"\bfear\s+(?:and|&)\s+greed\b",)),
    ("rating change", (r"\brating\s+(?:change|upgrade|downgrade)\b",)),
    ("recommendation change", (r"\b(?:recommendation|upgrade|downgrade|initiated)\b",)),
    ("guidance", (r"\bguidance\b",)),
    ("estimate revision", (r"\bestimate\s+(?:revision|change|increase|decrease)\b",)),
]

FINANCE_KEYWORDS = {
    "earnings", "revenue", "margin", "guidance", "estimate", "credit", "rating",
    "dividend", "liquidity", "leverage", "capital", "cash flow", "valuation",
    "outlook", "quarter", "results", "transcript", "10-q", "8-k", "research",
    "recommendation", "sector", "upgrade", "downgrade",
}

TICKER_STOPWORDS = {
    "PDF", "XLSX", "XLS", "CSV", "TXT", "MD", "Q", "FY", "FQ", "USD", "US",
    "USA", "UK", "EU", "SEC", "CEO", "CFO", "EPS", "EBITDA", "NAV", "REIT",
    "JPM", "GS", "MS", "RBC", "WF", "BMO", "UBS", "CITI", "BOFA", "BAML", "MD&A",
    "HOVDE", "BREAN", "WOLFE", "BAIRD", "PIPER",
    "BANKS", "COMPANIES", "CREDIT", "RESEARCH", "ROOT", "GREEN", "STREET",
}

DOCUMENT_TYPE_SCORES = {
    "earnings release": 28,
    "earnings presentation": 27,
    "earnings supplement": 24,
    "transcript": 28,
    "10-Q/10Q": 28,
    "8-K/8K": 25,
    "current report": 18,
    "credit report": 30,
    "analyst report": 27,
    "sector report": 26,
    "Green Street report": 28,
    "industry report": 18,
    "prepared remarks": 22,
    "guidance": 28,
    "estimate revision": 27,
    "recommendation change": 29,
    "rating change": 28,
    "press release": 14,
    "data supplement": 15,
    "MD&A": 20,
    "company overview": 12,
    "dividend announcement": 13,
    "insider trading": 12,
    "fear and greed": 8,
    "other": 4,
}

INVENTORY_COLUMNS = [
    "category", "ticker", "all_detected_tickers", "company_or_identifier", "source_or_broker", "document_type",
    "file_name", "file_extension", "file_size_mb", "relative_path", "modified_date",
    "extraction_status", "extracted_path",
]
RELEVANCE_COLUMNS = INVENTORY_COLUMNS + [
    "relevance_score", "priority_level", "reason_for_score", "score_breakdown",
    "investment_rationale",
]
CROSS_DAY_CHANGE_COLUMNS = [
    "change_type", "ticker", "category", "source_or_broker", "document_type",
    "file_name", "previous_score", "current_score", "priority_level", "reason_for_change",
]
CROSS_DAY_TICKER_COLUMNS = [
    "ticker", "previous_file_count", "current_file_count", "new_file_count",
    "removed_file_count", "new_brokers_or_sources", "new_document_types",
    "new_credit_report_count", "new_earnings_file_count", "new_transcript_count",
    "new_filing_count", "current_high_priority_count", "attention_reason",
]
RESEARCH_INDEX_COLUMNS = [
    "source_date", "indexed_source", "ticker", "all_detected_tickers",
    "source_or_broker", "category", "document_type", "file_name", "relative_path",
    "relevance_score", "priority_level", "extracted_snippet", "evidence_type",
    "extraction_status",
]
RESEARCH_SEARCH_RESULT_COLUMNS = RESEARCH_INDEX_COLUMNS + [
    "search_score", "match_type", "evidence_strength", "why_matched",
    "needs_manual_review", "matched_keywords", "display_snippet", "match_reason",
]
DASHBOARD_SIGNAL_COLUMNS = [
    "ticker", "company_or_identifier", "signal_type", "signal_direction", "confidence",
    "priority", "evidence", "source_file", "broker_or_source", "source_date", "category",
    "document_type", "why_it_matters", "why_review", "needs_manual_review",
]
DASHBOARD_PRIORITY_COLUMNS = [
    "ticker", "attention_score", "file_count", "high_priority_count", "broker_source_count",
    "credit_report_count", "earnings_transcript_filing_count", "categories_present",
    "attention_reason",
]
DASHBOARD_MANUAL_REVIEW_COLUMNS = [
    "ticker", "priority", "signal_type", "signal_direction", "confidence", "evidence",
    "source_file", "broker_or_source", "why_it_matters", "why_review",
]

DASHBOARD_SIGNAL_PATTERNS = [
    ("Rating / Price Target", (r"\brating\b", r"\bprice target\b", r"\bpt\b", r"\boverweight\b", r"\bunderweight\b", r"\bdowngrade\b", r"\bupgrade\b", r"\boutperform\b", r"\bunderperform\b")),
    ("Earnings / Results", (r"\bearnings?\b", r"\beps\b", r"\bresults?\b", r"\bquarter\b", r"\bbeat\b", r"\bmiss\b")),
    ("Guidance / Estimates", (r"\bguidance\b", r"\bestimates?\b", r"\brevision\b", r"\bforecast\b")),
    ("Credit / Liquidity", (r"\bcredit\b", r"\bliquidity\b", r"\bleverage\b", r"\bdebt\b", r"\bcovenant\b", r"\bissuance\b")),
    ("Valuation / Thesis", (r"\bvaluation\b", r"\bthesis\b", r"\bupside\b", r"\bdownside\b", r"\bmultiple\b")),
    ("Margin / Operating Performance", (r"\bmargins?\b", r"\boperating income\b", r"\boperating performance\b")),
    ("Capex / Investment Spend", (r"\bcapex\b", r"\bcapital expenditures?\b", r"\binvestment spend\b")),
    ("Cash Flow / FCF", (r"\bfree cash flow\b", r"\bfcf\b", r"\bcash flow\b")),
    ("Revenue / Growth", (r"\brevenue\b", r"\bsales\b", r"\bgrowth\b", r"\btop line\b")),
    ("Risk / Outlook", (r"\brisk\b", r"\boutlook\b", r"\bconcern\b")),
    ("M&A / Strategic Action", (r"\bmerger\b", r"\bacquisition\b", r"\bstrategic review\b", r"\basset sale\b", r"\bspin-?off\b", r"\bdivestiture\b")),
]

DASHBOARD_SIGNAL_TYPE_RANK = {
    "Rating / Price Target": 0,
    "Guidance / Estimates": 1,
    "Credit / Liquidity": 2,
    "Earnings / Results": 3,
    "Valuation / Thesis": 4,
    "Margin / Operating Performance": 5,
    "Capex / Investment Spend": 6,
    "Cash Flow / FCF": 7,
    "Revenue / Growth": 8,
    "M&A / Strategic Action": 9,
    "Risk / Outlook": 10,
    "Green Street / Sector": 11,
    "Filing / Transcript": 12,
    "Broker Coverage": 13,
    "Other": 14,
}

DASHBOARD_DIRECTION_PATTERNS = {
    "Rating / Price Target": {
        "Positive": (
            r"\bupgrade[sd]?\b", r"\braise[sd]?\s+(?:the\s+)?(?:rating|price target|pt)\b",
            r"\b(?:rating|price target|pt)\s+(?:was\s+)?raise[sd]?\b",
            r"\binitiat(?:e|es|ed)\s+(?:at|with)\s+(?:an?\s+)?(?:overweight|outperform|buy)\b",
            r"\b(?:overweight|outperform)\b",
        ),
        "Negative": (
            r"\bdowngrade[sd]?\b", r"\blower(?:ed|s)?\s+(?:the\s+)?(?:rating|price target|pt)\b",
            r"\bcut(?:s)?\s+(?:the\s+)?(?:rating|price target|pt)\b",
            r"\b(?:rating|price target|pt)\s+(?:was\s+)?lowered\b",
            r"\binitiat(?:e|es|ed)\s+(?:at|with)\s+(?:an?\s+)?(?:underweight|underperform|sell)\b",
            r"\b(?:underweight|underperform)\b",
        ),
    },
    "Earnings / Results": {
        "Positive": (r"\bbeat(?:s|en)?\b", r"\babove expectations?\b", r"\bstrong (?:results?|quarter)\b"),
        "Negative": (r"\bmiss(?:es|ed)?\b", r"\bbelow expectations?\b", r"\bweak(?:er)? (?:results?|quarter)\b"),
    },
    "Guidance / Estimates": {
        "Positive": (
            r"\braise[sd]?\s+(?:its\s+|the\s+)?(?:guidance|estimates?|forecast)\b",
            r"\b(?:guidance|estimates?|forecast)\s+(?:was\s+|were\s+)?raise[sd]?\b",
        ),
        "Negative": (
            r"\b(?:lower(?:ed|s)?|cut(?:s)?)\s+(?:its\s+|the\s+)?(?:guidance|estimates?|forecast)\b",
            r"\b(?:guidance|estimates?|forecast)\s+(?:was\s+|were\s+)?(?:lowered|cut)\b",
        ),
    },
    "Credit / Liquidity": {
        "Positive": (
            r"\bupgrade[sd]?\b", r"\bliquidity improve[sd]?\b", r"\bdeleverag(?:e|ed|ing)\b",
            r"\bleverage (?:declined|decreased|improved)\b", r"\bcredit improve[sd]?\b",
        ),
        "Negative": (
            r"\bdowngrade[sd]?\b", r"\bliquidity (?:pressure|concern|weakened)\b",
            r"\bleverage (?:concern|increased|rose|pressure)\b", r"\bissuance (?:need|required|pressure)\b",
            r"\bdebt (?:concern|pressure)\b", r"\bcovenant (?:risk|breach|concern)\b",
            r"\bcredit deterioration\b", r"\brating pressure\b",
        ),
    },
    "Capex / Investment Spend": {
        "Positive": (r"\bcapex (?:declined|decreased|improved)\b",),
        "Negative": (r"\b(?:rising|higher|increased) capex\b", r"\bcapex pressure\b", r"\bcapex .*hurt(?:s|ing)? (?:free cash flow|fcf)\b"),
    },
    "Cash Flow / FCF": {
        "Positive": (r"\b(?:free cash flow|fcf) improve[sd]?\b", r"\bstronger (?:free cash flow|fcf)\b"),
        "Negative": (r"\bnegative (?:free cash flow|fcf)\b", r"\b(?:free cash flow|fcf) pressure\b", r"\bweaker (?:free cash flow|fcf)\b"),
    },
    "Revenue / Growth": {
        "Positive": (r"\braise[sd]? revenue\b", r"\bstronger growth\b", r"\b(?:revenue|sales) beat\b"),
        "Negative": (r"\bgrowth slowdown\b", r"\b(?:revenue|sales) miss\b", r"\bweaker growth\b"),
    },
}

DASHBOARD_WHY_IT_MATTERS = {
    "Rating / Price Target": "May affect how broker positioning is interpreted.",
    "Earnings / Results": "May affect assessment of reported operating performance.",
    "Guidance / Estimates": "May affect forward estimates.",
    "Credit / Liquidity": "Could affect credit risk perception.",
    "Valuation / Thesis": "May affect valuation assumptions or thesis review.",
    "Margin / Operating Performance": "May affect operating-performance expectations.",
    "Capex / Investment Spend": "May affect investment-spend expectations.",
    "Cash Flow / FCF": "May impact free cash flow expectations.",
    "Revenue / Growth": "May affect revenue or growth expectations.",
    "Risk / Outlook": "Highlights an area for risk review.",
    "Broker Coverage": "Helps prioritize broker research review.",
    "Green Street / Sector": "May affect sector-level research prioritization.",
    "Filing / Transcript": "Flags primary-source material for review.",
    "M&A / Strategic Action": "May affect strategic-action review.",
    "Other": "Provides additional research context.",
}

INSUFFICIENT_TEXT_NOTICE = (
    "No verified financial claims were generated because extracted text was insufficient. "
    "The observations below are based on filenames, folders, document types, and "
    "broker/source coverage only."
)
SENSITIVE_FINANCE_PATTERNS = [
    r"\bbeat(?:s|en|ing)?\b", r"\bmiss(?:es|ed|ing)?\b", r"\beps\b", r"\brevenue\b",
    r"\bprice target\b", r"\brating (?:was |has been )?(?:raised|lowered|upgraded|downgraded)\b",
    r"\brating (?:changed|moved|cut|increased|decreased)\b",
    r"\bguidance (?:was |has been )?(?:raised|lowered|increased|decreased|cut)\b",
    r"\bunderwriting performance\b", r"\bcredit (?:deterioration|improvement)\b",
    r"\bcredit (?:improved|deteriorated|weakened|strengthened)\b",
    r"\bmargins?\b", r"\bcapex\b", r"\bliquidity\b", r"\bvaluation\b",
    r"\b(?:market weight|overweight|underweight)\b",
    r"\boperational (?:strength|weakness|performance)\b", r"\bperformance was (?:robust|strong|weak)\b",
    r"\bstrong earnings\b", r"\bweak earnings\b",
]

BROKER_COMPARISON_CATEGORIES = {"Research", "Banks", "Credit", "Companies", "Green Street"}
BROKER_COMPARISON_DOCUMENT_TYPES = {
    "analyst report", "credit report", "Green Street report", "sector report",
    "industry report", "recommendation change", "rating change", "guidance",
    "estimate revision",
}
COMPARISON_REVIEW_TERMS = {
    "guidance": r"\bguidance\b",
    "rating": r"\brating\b",
    "estimate": r"\bestimates?\b",
    "margin": r"\bmargins?\b",
    "capex": r"\bcapex\b|\bcapital expenditures?\b",
    "liquidity": r"\bliquidity\b",
    "valuation": r"\bvaluation\b",
    "credit": r"\bcredit\b",
}
DISCLOSURE_PHRASES = [
    "intended for institutional investors", "independence and disclosure standards",
    "finra", "conflicts of interest", "seeks to do business", "analyst certification",
    "important disclosures", "investors should consider this report", "all rights reserved",
    "use this report only", "not subject to", "discretionary basis",
    "covered in its research reports",
]
INVESTMENT_EVIDENCE_PATTERNS = {
    "Rating / PT": (
        r"\brating\b", r"\bprice target\b", r"\boverweight\b", r"\bunderweight\b",
        r"\bmarket weight\b", r"\bneutral\b", r"\bbuy\b", r"\bsell\b", r"\boutperform\b",
    ),
    "Earnings / Results": (
        r"\beps\b", r"\brevenue\b", r"\bsales\b", r"\bmargin\b", r"\boperating income\b",
    ),
    "Guidance / Estimates": (r"\bguidance\b", r"\bestimates?\b", r"\brevision\b"),
    "Credit / Liquidity": (r"\bliquidity\b", r"\bleverage\b", r"\bdebt\b", r"\bcredit\b"),
    "Valuation / Thesis": (r"\bvaluation\b", r"\bthesis\b", r"\bupside\b", r"\bdownside\b"),
    "Risk / Outlook": (r"\brisk\b", r"\boutlook\b"),
    "Other": (r"\bcapex\b", r"\baws\b", r"\bcloud\b", r"\bretail\b", r"\badvertising\b"),
}


def create_session_dir() -> Path:
    return Path(tempfile.mkdtemp(prefix="cutler_daily_research_"))


def parse_daily_folder_date_from_name(source_name: str) -> Optional[datetime]:
    """Parse a daily research date from an archive or source-folder name."""
    name = PurePosixPath(str(source_name or "").replace("\\", "/")).name.strip()
    name = re.sub(r"\.(?:zip|pdf|xlsx?|csv|txt|md)$", "", name, flags=re.IGNORECASE)
    for match in re.finditer(r"(?<!\d)(\d{1,2})[._-](\d{1,2})[._-](\d{2,4})(?!\d)", name):
        month, day, year = (int(value) for value in match.groups())
        if year < 100:
            year += 2000
        try:
            return datetime(year, month, day)
        except ValueError:
            continue
    return None


def daily_source_title_suffix(source_name: str, relative_paths: Optional[Iterable[str]] = None) -> str:
    candidates = [str(source_name or "")]
    for relative_path in relative_paths if relative_paths is not None else []:
        parts = PurePosixPath(str(relative_path).replace("\\", "/")).parts
        if parts:
            candidates.append(parts[0])
    for candidate in candidates:
        parsed = parse_daily_folder_date_from_name(candidate)
        if parsed:
            return parsed.strftime("%B %d, %Y").replace(" 0", " ")
    stem = Path(str(source_name or "daily_research")).stem.strip() or "daily_research"
    return f"Source Folder: {stem}"


def add_generation_method(markdown: str, method: str) -> str:
    """Insert generation metadata immediately below a Markdown title."""
    lines = str(markdown or "").splitlines()
    marker = f"Generation method: {method}"
    if any(line.strip() == marker for line in lines[:6]):
        return markdown
    if lines and lines[0].startswith("#"):
        return "\n".join([lines[0], "", marker, *lines[1:]])
    return f"{marker}\n\n{markdown}"


def remove_session_dir(path: str | Path | None) -> None:
    if not path:
        return
    try:
        target = Path(path).resolve()
        if target.exists() and target.name.startswith("cutler_daily_research_"):
            shutil.rmtree(target)
    except Exception:
        pass


def _is_junk(parts: Iterable[str], name: str) -> bool:
    lowered = {p.lower() for p in parts}
    return (
        name.lower() in JUNK_NAMES
        or name.startswith("._")
        or bool(lowered & JUNK_PARTS)
    )


def _safe_member_path(root: Path, member_name: str) -> Optional[Path]:
    normalized = member_name.replace("\\", "/")
    pure = PurePosixPath(normalized)
    if pure.is_absolute() or ".." in pure.parts:
        return None
    target = (root / Path(*pure.parts)).resolve()
    try:
        target.relative_to(root.resolve())
    except ValueError:
        return None
    return target


def safe_extract_zip(
    zip_path: str | Path,
    extract_root: str | Path,
    *,
    max_total_uncompressed: int = 2 * 1024 * 1024 * 1024,
    max_member_size: int = 500 * 1024 * 1024,
) -> Tuple[List[Dict[str, Any]], List[str]]:
    root = Path(extract_root)
    root.mkdir(parents=True, exist_ok=True)
    records: List[Dict[str, Any]] = []
    warnings: List[str] = []
    extracted_total = 0

    with zipfile.ZipFile(zip_path) as archive:
        for info in archive.infolist():
            if info.is_dir():
                continue
            rel = info.filename.replace("\\", "/")
            pure = PurePosixPath(rel)
            base = pure.name
            status = "extracted"
            target = _safe_member_path(root, rel)

            if target is None:
                status = "skipped_unsafe_path"
            elif _is_junk(pure.parts, base):
                status = "skipped_junk"
            elif info.file_size > max_member_size:
                status = "skipped_too_large"
            elif extracted_total + info.file_size > max_total_uncompressed:
                status = "skipped_total_size_limit"

            if status == "extracted":
                try:
                    target.parent.mkdir(parents=True, exist_ok=True)
                    with archive.open(info) as src, target.open("wb") as dst:
                        shutil.copyfileobj(src, dst, length=1024 * 1024)
                    extracted_total += info.file_size
                except Exception as exc:
                    status = f"failed: {type(exc).__name__}"
                    warnings.append(f"{rel}: {exc}")

            records.append(
                {
                    "relative_path": rel,
                    "file_name": base,
                    "file_size": int(info.file_size),
                    "modified_date": _zip_modified_date(info),
                    "extraction_status": status,
                    "extracted_path": str(target) if target and status == "extracted" else "",
                }
            )
    return records, warnings


def _zip_modified_date(info: zipfile.ZipInfo) -> str:
    try:
        return datetime(*info.date_time).isoformat(timespec="seconds")
    except Exception:
        return ""


def detect_category(relative_path: str) -> str:
    parts = [p.lower() for p in PurePosixPath(relative_path.replace("\\", "/")).parts[:-1]]
    for category in CATEGORIES:
        label = category.lower()
        if label in parts or (category == "Green Street" and {"green", "street"} <= set(parts)):
            return category
    joined = " ".join(parts)
    for category in CATEGORIES:
        if category.lower() in joined:
            return category
    return "Root"


def detect_source(text: str) -> str:
    normalized = _normalize_metadata_text(text)
    for label, pattern in BROKER_PATTERNS:
        if re.search(pattern, normalized, flags=re.IGNORECASE):
            return label
    return ""


def _normalize_metadata_text(text: str) -> str:
    normalized = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", text)
    normalized = re.sub(r"(?<=[A-Z])(?=[A-Z][a-z])", " ", normalized)
    normalized = re.sub(r"[_\-.]+", " ", normalized)
    return re.sub(r"\s+", " ", normalized).strip()


def detect_document_type(text: str, *, category: str = "Root", source: str = "") -> str:
    normalized = _normalize_metadata_text(text)
    for label, patterns in DOCUMENT_TYPES:
        if any(re.search(pattern, normalized, flags=re.IGNORECASE) for pattern in patterns):
            return label
    if category == "Credit" and (source or re.search(r"\bcredit\b", normalized, flags=re.IGNORECASE)):
        return "credit report"
    if category == "Green Street" or source == "Green Street":
        return "Green Street report" if source == "Green Street" else "sector report"
    if source and category in {"Research", "Banks", "Companies", "Root"}:
        return "analyst report"
    return "other"


def detect_ticker(
    file_name: str,
    known_tickers: Optional[set[str]] = None,
    *,
    source: str = "",
    category: str = "Root",
) -> str:
    stem = Path(file_name).stem
    tokens = re.findall(r"(?<![A-Za-z0-9])[A-Z]{1,5}(?![A-Za-z0-9])", stem)
    candidates = [t for t in tokens if t not in TICKER_STOPWORDS and not re.fullmatch(r"Q[1-4]", t)]
    if known_tickers:
        known = [t for t in candidates if t in known_tickers]
        if known:
            return known[0]
    if not candidates:
        return ""
    if source or category in {"Banks", "Companies", "Credit", "Research"}:
        # A recognized broker following a leading all-caps token is a strong,
        # conservative research-filename convention even when the ticker universe is stale.
        first_token = re.match(r"^\s*([A-Z]{1,5})(?=\s|[_\-.])", stem)
        if first_token and first_token.group(1) in candidates:
            return first_token.group(1)
    if known_tickers:
        return ""
    first = candidates[0]
    return first if stem.upper().count(first) >= 2 else ""


def detect_all_tickers(
    file_name: str,
    known_tickers: Optional[set[str]] = None,
    *,
    source: str = "",
    category: str = "Root",
) -> List[str]:
    stem = Path(file_name).stem
    tokens = re.findall(r"(?<![A-Za-z0-9])[A-Z]{1,5}(?![A-Za-z0-9])", stem)
    candidates = [t for t in tokens if t not in TICKER_STOPWORDS and not re.fullmatch(r"Q[1-4]", t)]
    detected: List[str] = []
    for token in candidates:
        if token in detected:
            continue
        if known_tickers and token in known_tickers:
            detected.append(token)
        elif source and token == candidates[0]:
            detected.append(token)
        elif category == "Credit" and len(candidates) > 1:
            detected.append(token)
    return detected


def company_or_identifier(file_name: str, ticker: str, source: str, document_type: str) -> str:
    stem = Path(file_name).stem
    cleaned = re.sub(r"[_\-.]+", " ", stem)
    for token in (ticker, source, document_type):
        if token:
            cleaned = re.sub(re.escape(token), " ", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\b(?:20\d{2}|19\d{2}|Q[1-4]|FY\d{2,4})\b", " ", cleaned, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", cleaned).strip()[:160]


def build_inventory(records: List[Dict[str, Any]], known_tickers: Optional[set[str]] = None) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    for record in records:
        rel = str(record.get("relative_path") or "")
        name = str(record.get("file_name") or Path(rel).name)
        category = detect_category(rel)
        source = detect_source(f"{rel} {name}")
        doc_type = detect_document_type(name, category=category, source=source)
        ticker = detect_ticker(name, known_tickers=known_tickers, source=source, category=category)
        all_tickers = detect_all_tickers(
            name,
            known_tickers=known_tickers,
            source=source,
            category=category,
        )
        if ticker and ticker not in all_tickers:
            all_tickers.insert(0, ticker)
        rows.append(
            {
                "category": category,
                "ticker": ticker,
                "all_detected_tickers": ", ".join(all_tickers),
                "company_or_identifier": company_or_identifier(name, ticker, source, doc_type),
                "source_or_broker": source,
                "document_type": doc_type,
                "file_name": name,
                "file_extension": Path(name).suffix.lower(),
                "file_size_mb": round(float(record.get("file_size") or 0) / (1024 * 1024), 3),
                "relative_path": rel,
                "modified_date": record.get("modified_date") or "",
                "extraction_status": record.get("extraction_status") or "",
                "extracted_path": record.get("extracted_path") or "",
            }
        )
    return pd.DataFrame(rows, columns=INVENTORY_COLUMNS)


def score_inventory(inventory: pd.DataFrame) -> pd.DataFrame:
    if inventory.empty:
        return pd.DataFrame(columns=RELEVANCE_COLUMNS)
    rows: List[Dict[str, Any]] = []
    ticker_counts = Counter(str(x) for x in inventory.get("ticker", []) if str(x).strip())
    ticker_categories: Dict[str, set[str]] = {}
    ticker_sources: Dict[str, set[str]] = {}
    for _, source_row in inventory.iterrows():
        tickers_in_row = [
            t.strip() for t in str(source_row.get("all_detected_tickers") or source_row.get("ticker") or "").split(",")
            if t.strip()
        ]
        for detected in tickers_in_row:
            ticker_categories.setdefault(detected, set()).add(str(source_row.get("category") or "Root"))
            source = str(source_row.get("source_or_broker") or "")
            if source:
                ticker_sources.setdefault(detected, set()).add(source)

    for _, row in inventory.iterrows():
        category = str(row.get("category") or "Root")
        ticker = str(row.get("ticker") or "")
        source = str(row.get("source_or_broker") or "")
        doc_type = str(row.get("document_type") or "other")
        name = str(row.get("file_name") or "")
        status = str(row.get("extraction_status") or "")
        searchable = re.sub(r"[_\-.]+", " ", f"{name} {row.get('relative_path') or ''}").lower()
        breakdown: List[str] = []
        score = 0

        category_points = {
            "Credit": 20, "Companies": 18, "Banks": 18, "Research": 16,
            "Green Street": 22, "Root": 8,
        }.get(category, 8)
        score += category_points
        breakdown.append(f"{category} folder +{category_points}")

        type_points = DOCUMENT_TYPE_SCORES.get(doc_type, 4)
        score += type_points
        breakdown.append(f"{doc_type} +{type_points}")

        if source:
            score += 8
            breakdown.append(f"recognized source {source} +8")
            if source in TOP_TIER_BROKERS:
                score += 5
                breakdown.append("top-tier broker +5")
            elif source in SPECIALTY_SOURCES:
                score += 5
                breakdown.append("specialty source +5")
        if ticker:
            score += 12
            breakdown.append(f"recognized ticker {ticker} +12")
            if ticker_counts[ticker] > 1:
                boost = min(10, (ticker_counts[ticker] - 1) * 2)
                score += boost
                breakdown.append(f"repeated ticker coverage +{boost}")
            if len(ticker_categories.get(ticker, set())) > 1:
                score += 6
                breakdown.append("cross-category ticker coverage +6")
            if len(ticker_sources.get(ticker, set())) > 1:
                boost = min(10, len(ticker_sources[ticker]) * 2)
                score += boost
                breakdown.append(f"multiple broker/source coverage +{boost}")

        keyword_hits = sorted({kw for kw in FINANCE_KEYWORDS if kw in searchable})
        if keyword_hits:
            boost = min(12, len(keyword_hits) * 3)
            score += boost
            breakdown.append(f"finance keywords ({', '.join(keyword_hits[:4])}) +{boost}")

        if status != "extracted":
            score = max(0, score - 40)
            breakdown.append("not extracted -40")
        if str(row.get("file_extension") or "") not in SUPPORTED_TEXT_EXTENSIONS:
            score = max(0, score - 8)
            breakdown.append("unsupported text type -8")

        score = min(100, int(score))
        priority = "High" if score >= 70 else "Medium" if score >= 45 else "Low"
        rationale_parts = [f"{doc_type} in {category}"]
        if ticker:
            rationale_parts.append(f"covers {ticker}")
        if source:
            rationale_parts.append(f"from {source}")
        if ticker and len(ticker_sources.get(ticker, set())) > 1:
            rationale_parts.append("with multiple-source coverage")
        if ticker and len(ticker_categories.get(ticker, set())) > 1:
            rationale_parts.append("appearing across categories")
        rationale = "; ".join(rationale_parts) + "."
        enriched = row.to_dict()
        enriched.update(
            {
                "relevance_score": score,
                "priority_level": priority,
                "reason_for_score": rationale,
                "score_breakdown": "; ".join(breakdown),
                "investment_rationale": rationale,
            }
        )
        rows.append(enriched)

    result = pd.DataFrame(rows)
    if not result.empty:
        result = result.sort_values(
            ["relevance_score", "ticker", "file_name"],
            ascending=[False, True, True],
        ).reset_index(drop=True)
    return result


def select_files_for_text(relevance: pd.DataFrame, max_files: int) -> pd.DataFrame:
    if relevance.empty:
        return relevance.copy()
    eligible = relevance[
        (relevance["extraction_status"] == "extracted")
        & (relevance["file_extension"].isin(SUPPORTED_TEXT_EXTENSIONS))
        & (relevance["priority_level"] == "High")
    ]
    return eligible.head(int(max_files)).copy()


def extract_selected_text(
    selected: pd.DataFrame,
    *,
    max_pdf_pages: int,
    max_chars_per_file: int,
) -> List[Dict[str, Any]]:
    output: List[Dict[str, Any]] = []
    for _, row in selected.iterrows():
        path = Path(str(row.get("extracted_path") or ""))
        text = ""
        status = "not_scanned"
        try:
            suffix = path.suffix.lower()
            if suffix == ".pdf":
                text = _extract_pdf(path, max_pdf_pages, max_chars_per_file)
            elif suffix == ".xlsx":
                text = _extract_xlsx_preview(path, max_chars_per_file)
            elif suffix == ".xls":
                text = "Legacy XLS file inventoried; lightweight preview is not available."
            elif suffix in {".txt", ".md", ".csv"}:
                text = path.read_text(encoding="utf-8", errors="replace")[:max_chars_per_file]
            status = "scanned" if text.strip() else "scanned_no_text"
        except Exception as exc:
            status = f"scan_failed: {type(exc).__name__}"
            text = ""
        output.append(
            {
                "relative_path": row.get("relative_path") or "",
                "file_name": row.get("file_name") or "",
                "ticker": row.get("ticker") or "",
                "category": row.get("category") or "",
                "source_or_broker": row.get("source_or_broker") or "",
                "document_type": row.get("document_type") or "",
                "relevance_score": int(row.get("relevance_score") or 0),
                "text_extraction_status": status,
                "extracted_text": text[:max_chars_per_file],
            }
        )
    return output


def _extract_pdf(path: Path, max_pages: int, max_chars: int) -> str:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    parts: List[str] = []
    for page in reader.pages[:max_pages]:
        parts.append(page.extract_text() or "")
        if sum(len(p) for p in parts) >= max_chars:
            break
    return "\n".join(parts)[:max_chars]


def _extract_xlsx_preview(path: Path, max_chars: int) -> str:
    try:
        from openpyxl import load_workbook
    except Exception:
        return "XLSX file inventoried; openpyxl is unavailable for preview."
    workbook = load_workbook(path, read_only=True, data_only=True)
    lines = ["Sheets: " + ", ".join(workbook.sheetnames)]
    for sheet in workbook.worksheets[:3]:
        lines.append(f"\n[{sheet.title}]")
        for row in sheet.iter_rows(min_row=1, max_row=8, values_only=True):
            values = [str(v) for v in row[:8] if v is not None]
            if values:
                lines.append(" | ".join(values))
            if len("\n".join(lines)) >= max_chars:
                break
    workbook.close()
    return "\n".join(lines)[:max_chars]


def build_ticker_summary(relevance: pd.DataFrame) -> pd.DataFrame:
    if relevance.empty:
        return pd.DataFrame()
    known = _expand_rows_by_detected_ticker(relevance)
    rows = []
    for ticker, group in known.groupby("ticker"):
        sources = sorted({str(x) for x in group["source_or_broker"] if str(x)})
        categories = sorted({str(x) for x in group["category"] if str(x)})
        doc_types = group["document_type"].astype(str)
        broker_reports = group[group["source_or_broker"].astype(str).str.strip() != ""]
        attention = _attention_reason(group, ticker=ticker)
        rows.append(
            {
                "ticker": ticker,
                "file_count": len(group),
                "high_priority_files": int((group["priority_level"] == "High").sum()),
                "max_relevance_score": int(group["relevance_score"].max()),
                "broker_report_count": len(broker_reports),
                "credit_report_count": int((doc_types == "credit report").sum()),
                "earnings_file_count": int(doc_types.str.startswith("earnings").sum()),
                "filing_file_count": int(doc_types.isin(["10-Q/10Q", "8-K/8K", "current report", "MD&A"]).sum()),
                "transcript_count": int((doc_types == "transcript").sum()),
                "has_multiple_brokers": len(sources) > 1,
                "categories": ", ".join(categories),
                "sources": ", ".join(sources),
                "attention_reason": attention,
            }
        )
    return pd.DataFrame(rows).sort_values(["max_relevance_score", "file_count"], ascending=False) if rows else pd.DataFrame()


def build_broker_coverage_summary(relevance: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "ticker", "broker_report_count", "brokers_or_sources", "analyst_report_files",
        "credit_report_files", "categories_present", "attention_reason",
    ]
    if relevance.empty:
        return pd.DataFrame(columns=columns)
    rows = []
    known = _expand_rows_by_detected_ticker(relevance)
    for ticker, group in known.groupby("ticker"):
        broker_group = group[group["source_or_broker"].astype(str).str.strip() != ""]
        if broker_group.empty:
            continue
        report_group = broker_group
        rows.append(
            {
                "ticker": ticker,
                "broker_report_count": len(report_group),
                "brokers_or_sources": ", ".join(sorted(set(report_group["source_or_broker"].astype(str)))),
                "analyst_report_files": int((report_group["document_type"] == "analyst report").sum()),
                "credit_report_files": int((report_group["document_type"] == "credit report").sum()),
                "categories_present": ", ".join(sorted(set(report_group["category"].astype(str)))),
                "attention_reason": _attention_reason(group, ticker=ticker),
            }
        )
    return (
        pd.DataFrame(rows, columns=columns).sort_values(
            ["broker_report_count", "ticker"], ascending=[False, True]
        )
        if rows else pd.DataFrame(columns=columns)
    )


def select_broker_comparison_files(
    relevance: pd.DataFrame,
    ticker: str,
    *,
    max_files: int,
) -> pd.DataFrame:
    """Select same-day broker/source reports for one ticker, ordered by relevance."""
    if relevance.empty or not ticker:
        return relevance.iloc[0:0].copy()
    expanded = _expand_rows_by_detected_ticker(relevance)
    ticker_rows = expanded[expanded["ticker"].astype(str).str.upper() == ticker.upper()].copy()
    if ticker_rows.empty:
        return ticker_rows
    broker_rows = ticker_rows[
        ticker_rows["source_or_broker"].astype(str).str.strip().ne("")
        & ticker_rows["category"].astype(str).isin(BROKER_COMPARISON_CATEGORIES)
    ].copy()
    preferred = broker_rows[
        broker_rows["document_type"].astype(str).isin(BROKER_COMPARISON_DOCUMENT_TYPES)
    ].copy()
    selected = preferred if not preferred.empty else broker_rows[
        ~broker_rows["document_type"].astype(str).isin(
            ["10-Q/10Q", "8-K/8K", "current report", "earnings release",
             "earnings presentation", "earnings supplement", "transcript", "prepared remarks"]
        )
    ].copy()
    if selected.empty:
        return selected
    selected = selected.drop_duplicates(subset=["relative_path"]).sort_values(
        ["relevance_score", "source_or_broker", "file_name"],
        ascending=[False, True, True],
    )
    return selected.head(int(max_files)).reset_index(drop=True)


def prepare_broker_comparison_text(
    selected: pd.DataFrame,
    existing_text: List[Dict[str, Any]],
    *,
    max_pdf_pages: int,
    max_chars_per_file: int,
    reuse_existing: bool,
) -> List[Dict[str, Any]]:
    """Reuse prior snippets when requested and scan only missing comparison files."""
    if selected.empty:
        return []
    existing_by_path = {
        str(item.get("relative_path") or ""): item
        for item in existing_text
        if str(item.get("relative_path") or "")
    }
    output: List[Dict[str, Any]] = []
    missing_rows = []
    for _, row in selected.iterrows():
        path = str(row.get("relative_path") or "")
        prior = existing_by_path.get(path)
        if reuse_existing and prior:
            reused = dict(prior)
            reused["extracted_text"] = str(reused.get("extracted_text") or "")[:max_chars_per_file]
            output.append(reused)
        else:
            missing_rows.append(row.to_dict())
    if missing_rows:
        missing = pd.DataFrame(missing_rows, columns=selected.columns)
        output.extend(
            extract_selected_text(
                missing,
                max_pdf_pages=max_pdf_pages,
                max_chars_per_file=max_chars_per_file,
            )
        )
    order = {str(row.get("relative_path") or ""): i for i, row in selected.iterrows()}
    return sorted(output, key=lambda item: order.get(str(item.get("relative_path") or ""), 10**9))


def _investment_evidence_type(text: str) -> str:
    scored = []
    for evidence_type, patterns in INVESTMENT_EVIDENCE_PATTERNS.items():
        hits = sum(bool(re.search(pattern, text, flags=re.IGNORECASE)) for pattern in patterns)
        if hits:
            scored.append((hits, evidence_type))
    return max(scored)[1] if scored else "Other"


def best_investment_useful_snippet(item: Dict[str, Any], *, max_chars: int = 500) -> Dict[str, Any]:
    """Choose an investment-useful extracted passage while avoiding disclosure-heavy text."""
    status = str(item.get("text_extraction_status") or "metadata-only")
    text = _normalized_extracted_text(item)
    result = {
        "evidence_type": "Other",
        "snippet": "",
        "quality": "unavailable",
        "extraction_status": status,
    }
    if status != "scanned" or not text:
        return result

    sentences = [
        sentence.strip()
        for sentence in re.split(r"(?<=[.!?])\s+|\s{2,}", text)
        if sentence.strip()
    ]
    if not sentences:
        sentences = [text]
    candidates = []

    def add_candidate(passage: str) -> None:
        passage = passage.strip()[:max_chars]
        if not passage:
            return
        useful_hits = sum(
            bool(re.search(pattern, passage, flags=re.IGNORECASE))
            for patterns in INVESTMENT_EVIDENCE_PATTERNS.values()
            for pattern in patterns
        )
        disclosure_hits = sum(phrase in passage.lower() for phrase in DISCLOSURE_PHRASES)
        score = (useful_hits * 6) - (disclosure_hits * 15)
        if useful_hits:
            score += min(len(passage), max_chars) / max_chars
        candidates.append((score, useful_hits, disclosure_hits, passage))

    for index in range(len(sentences)):
        for window_size in (1, 2, 3):
            add_candidate(" ".join(sentences[index:index + window_size]))
    for start in range(0, len(text), max(200, max_chars // 2)):
        add_candidate(text[start:start + max_chars])

    best = max(candidates, default=None, key=lambda value: value[0])
    if best and best[1] > 0 and best[0] > 0:
        passage = best[3]
        return {
            "evidence_type": _investment_evidence_type(passage),
            "snippet": passage,
            "quality": "investment_useful",
            "extraction_status": status,
        }
    if any(phrase in text.lower() for phrase in DISCLOSURE_PHRASES):
        result["quality"] = "disclosure_only"
    else:
        result["quality"] = "no_useful_snippet"
    return result


def _markdown_table_text(value: Any) -> str:
    return str(value or "").replace("|", r"\|").replace("\n", " ").strip()


def build_deterministic_broker_comparison(
    ticker: str,
    files: pd.DataFrame,
    snippets: List[Dict[str, Any]],
    *,
    report_date: str,
    mode: str,
) -> str:
    """Build a conservative same-day comparison from metadata and direct excerpts."""
    snippets_by_path = {str(item.get("relative_path") or ""): item for item in snippets}
    verified = verified_extracted_text_items(snippets)
    sources = sorted({str(x) for x in files["source_or_broker"] if str(x)}) if not files.empty else []
    lines = [
        f"# Broker Consensus Report - {ticker} - {report_date}",
        "",
        f"Comparison mode: {mode}",
        "",
        "## Files Compared",
        "",
        "| Broker/source | File name | Document type | Extraction status | Relevance score |",
        "|---|---|---|---|---:|",
    ]
    for _, row in files.iterrows():
        path = str(row.get("relative_path") or "")
        item = snippets_by_path.get(path, {})
        status = str(item.get("text_extraction_status") or "metadata-only")
        lines.append(
            f"| {row.get('source_or_broker') or 'Unknown'} | `{row.get('file_name') or path}` | "
            f"{row.get('document_type') or 'other'} | {status} | {int(row.get('relevance_score') or 0)} |"
        )

    lines.extend(
        [
            "",
            "## Key Extracted Evidence",
            "",
            "| Broker/source | File name | Evidence type | Extracted evidence | Extraction status |",
            "|---|---|---|---|---|",
        ]
    )
    for _, row in files.iterrows():
        path = str(row.get("relative_path") or "")
        item = snippets_by_path.get(path, {})
        evidence = best_investment_useful_snippet(item)
        snippet = evidence["snippet"] or "No investment-useful snippet found in limited extraction."
        lines.append(
            f"| {_markdown_table_text(row.get('source_or_broker') or 'Unknown')} | "
            f"`{_markdown_table_text(row.get('file_name') or path)}` | "
            f"{evidence['evidence_type']} | \"{_markdown_table_text(snippet)}\" | "
            f"{_markdown_table_text(evidence['extraction_status'])} |"
        )

    lines.extend(["", "## Broker-by-Broker Summary"])
    if files.empty:
        lines.append("No broker/source reports available.")
    for broker, group in files.groupby("source_or_broker", sort=True):
        lines.append(f"### {broker}")
        for _, row in group.iterrows():
            path = str(row.get("relative_path") or "")
            item = snippets_by_path.get(path, {})
            status = str(item.get("text_extraction_status") or "metadata-only")
            evidence = best_investment_useful_snippet(item)
            lines.append(
                f"- `{path}`: detected as {row.get('document_type') or 'other'} in "
                f"{row.get('category') or 'Unknown'}; extraction status: {status}."
            )
            if evidence["snippet"]:
                lines.append(
                    f"  Investment-useful evidence is available in the Key Extracted Evidence table. "
                    f"Source: `{path}`"
                )
            elif evidence["quality"] == "disclosure_only":
                lines.append(
                    f"  Limited extraction mostly captured disclosure/disclaimer text; review PDF directly. "
                    f"Source: `{path}`"
                )
            else:
                lines.append(
                    f"  No investment-useful snippet found in limited extraction. Source: `{path}`"
                )

    lines.extend(["", "## Consensus Themes"])
    if len(sources) >= 2:
        cited = ", ".join(f"`{path}`" for path in files["relative_path"].astype(str).head(12))
        lines.append(
            f"- Multiple brokers/sources covered **{ticker}** in the uploaded research set: "
            f"{', '.join(sources)}. Sources: {cited}"
        )
    doc_counts = files["document_type"].astype(str).value_counts() if not files.empty else pd.Series(dtype=int)
    for document_type, count in doc_counts.items():
        if count < 2:
            continue
        paths = files.loc[files["document_type"].astype(str) == document_type, "relative_path"].astype(str)
        lines.append(
            f"- {count} files were detected as **{document_type}**. Sources: "
            + ", ".join(f"`{path}`" for path in paths)
        )
    verified_term_sources: Dict[str, List[str]] = {}
    for term, pattern in COMPARISON_REVIEW_TERMS.items():
        matched = [
            str(item.get("relative_path") or "")
            for item in verified
            if re.search(
                pattern,
                str(best_investment_useful_snippet(item).get("snippet") or ""),
                flags=re.IGNORECASE,
            )
        ]
        if len(set(matched)) >= 2:
            verified_term_sources[term] = sorted(set(matched))
            lines.append(
                f"- The term **{term}** appears in limited extracted text from at least two sources: "
                + ", ".join(f"`{path}`" for path in sorted(set(matched)))
                + ". This records term overlap only, not agreement or direction."
            )
    if len(sources) < 2 and not any(count >= 2 for count in doc_counts) and not verified_term_sources:
        lines.append("No consensus theme could be verified beyond the available file metadata.")

    lines.extend(["", "## Divergences / Differences"])
    lines.append("Not enough extracted text to verify broker-level differences.")

    lines.extend(
        [
            "",
            "## Items to Verify",
            "- Rating or price target changes in the source PDFs.",
            "- EPS or revenue beats/misses in the source PDFs.",
            "- Guidance changes and estimate revisions.",
            "- Credit concerns, liquidity commentary, and valuation assumptions.",
            "",
            "## Source References",
        ]
    )
    for _, row in files.iterrows():
        path = str(row.get("relative_path") or "")
        status = str(snippets_by_path.get(path, {}).get("text_extraction_status") or "metadata-only")
        lines.append(f"- `{path}` ({row.get('source_or_broker') or 'source unknown'}; {status})")
    return "\n".join(lines)


def build_broker_comparison_evidence_payload(
    ticker: str,
    files: pd.DataFrame,
    snippets: List[Dict[str, Any]],
    *,
    max_total_chars: int = 30000,
) -> str:
    snippets_by_path = {str(item.get("relative_path") or ""): item for item in snippets}
    records = []
    used = 0
    for _, row in files.iterrows():
        path = str(row.get("relative_path") or "")
        item = snippets_by_path.get(path, {})
        text = _normalized_extracted_text(item)
        status = str(item.get("text_extraction_status") or "metadata-only")
        verified = status == "scanned" and len(text) >= 160
        evidence = best_investment_useful_snippet(item, max_chars=900)
        record = {
            "relative_path": path,
            "file_name": row.get("file_name"),
            "ticker": ticker,
            "category": row.get("category"),
            "source_or_broker": row.get("source_or_broker"),
            "document_type": row.get("document_type"),
            "relevance_score": int(row.get("relevance_score") or 0),
            "investment_rationale": row.get("investment_rationale"),
            "text_extraction_status": status,
            "verified_text_available": verified,
            "evidence_type": evidence["evidence_type"],
            "evidence_quality": evidence["quality"],
            "limited_text": evidence["snippet"] if verified and evidence["snippet"] else INSUFFICIENT_TEXT_NOTICE,
        }
        encoded = json.dumps(record, ensure_ascii=False)
        if used + len(encoded) > max_total_chars:
            break
        records.append(record)
        used += len(encoded)
    return json.dumps(
        {
            "ticker": ticker,
            "grounding_rule": (
                "Metadata supports file/source/category/document-type observations only. Financial or broker-view "
                "claims require explicit verified limited text and a same-line source filename citation."
            ),
            "files": records,
        },
        ensure_ascii=False,
    )


def validate_broker_comparison_grounding(text: str, snippets: List[Dict[str, Any]]) -> bool:
    required = [
        "## Files Compared", "## Key Extracted Evidence", "## Broker-by-Broker Summary",
        "## Consensus Themes", "## Divergences / Differences", "## Items to Verify",
        "## Source References",
    ]
    if not all(section in text for section in required):
        return False
    key_section = text.split("## Key Extracted Evidence", 1)[1].split("## Broker-by-Broker Summary", 1)[0]
    verified = verified_extracted_text_items(snippets)
    for line in key_section.splitlines():
        stripped = line.strip()
        if not stripped.startswith("|") or "broker/source" in stripped.lower() or bool(re.fullmatch(r"[\s|:-]+", stripped)):
            continue
        cited = [
            item for item in snippets
            if str(item.get("relative_path") or "") in line or str(item.get("file_name") or "") in line
        ]
        if not cited:
            return False
        quotes = [part.strip() for part in re.findall(r'"([^"]+)"', line) if part.strip()]
        for quote in quotes:
            if quote == "No investment-useful snippet found in limited extraction.":
                continue
            if not any(quote.lower() in _normalized_extracted_text(item).lower() for item in cited):
                return False
    before_evidence, after_evidence = text.split("## Key Extracted Evidence", 1)
    after_evidence = after_evidence.split("## Broker-by-Broker Summary", 1)[1]
    claim_text = before_evidence + "## Broker-by-Broker Summary" + after_evidence
    claim_text = claim_text.split("## Items to Verify", 1)[0]
    if not validate_llm_brief_grounding(claim_text, snippets):
        return False
    source_names = {
        value
        for item in snippets
        for value in [str(item.get("relative_path") or ""), str(item.get("file_name") or "")]
        if value
    }
    for line in claim_text.splitlines():
        stripped = line.strip()
        if (
            not stripped
            or stripped.startswith("#")
            or bool(re.fullmatch(r"[\s|:-]+", stripped))
            or (stripped.startswith("|") and "broker" in stripped.lower() and "file" in stripped.lower())
            or stripped.startswith("Comparison mode:")
            or stripped == "Not enough extracted text to verify broker-level differences."
        ):
            continue
        if not any(source in stripped for source in source_names):
            return False
    divergence = text.split("## Divergences / Differences", 1)[1].split("## Items to Verify", 1)[0]
    if "Not enough extracted text to verify broker-level differences." not in divergence:
        cited_sources = {
            str(item.get("relative_path") or "")
            for item in verified
            if str(item.get("relative_path") or "") in divergence
        }
        matched_quotes = {
            quote
            for quote in re.findall(r'"([^"]+)"', divergence)
            if len(quote.strip()) >= 20
            and any(quote.lower() in _normalized_extracted_text(item).lower() for item in verified)
        }
        if len(cited_sources) < 2 or len(matched_quotes) < 2:
            return False
    return True


def select_ticker_memo_files(
    relevance: pd.DataFrame,
    ticker: str,
    *,
    max_files: int,
) -> pd.DataFrame:
    """Select a capped, diversified set of same-day documents for one ticker."""
    if relevance.empty or not ticker:
        return relevance.iloc[0:0].copy()
    expanded = _expand_rows_by_detected_ticker(relevance)
    rows = expanded[
        (expanded["ticker"].astype(str).str.upper() == ticker.upper())
    ].drop_duplicates(subset=["relative_path"]).copy()
    if rows.empty:
        return rows

    rows["_memo_type_rank"] = rows["document_type"].astype(str).map(
        {
            "credit report": 0, "earnings release": 0, "earnings presentation": 0,
            "transcript": 0, "10-Q/10Q": 0, "8-K/8K": 0, "analyst report": 1,
            "earnings supplement": 1, "prepared remarks": 1, "guidance": 1,
            "estimate revision": 1, "rating change": 1, "recommendation change": 1,
        }
    ).fillna(2)
    rows = rows.sort_values(
        ["_memo_type_rank", "relevance_score", "source_or_broker", "file_name"],
        ascending=[True, False, True, True],
    )
    selected_indices = []
    for _, group in rows.groupby("document_type", sort=False):
        selected_indices.append(group.index[0])
        if len(selected_indices) >= int(max_files):
            break
    for index in rows.index:
        if len(selected_indices) >= int(max_files):
            break
        if index not in selected_indices:
            selected_indices.append(index)
    return rows.loc[selected_indices].drop(columns=["_memo_type_rank"]).reset_index(drop=True)


def build_ticker_memo_evidence_rows(
    files: pd.DataFrame,
    snippets: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    snippets_by_path = {str(item.get("relative_path") or ""): item for item in snippets}
    rows = []
    for _, row in files.iterrows():
        path = str(row.get("relative_path") or "")
        item = snippets_by_path.get(path, {})
        evidence = best_investment_useful_snippet(item)
        rows.append(
            {
                "source_or_broker": row.get("source_or_broker") or "Company / filing source",
                "file_name": row.get("file_name") or path,
                "document_type": row.get("document_type") or "other",
                "category": row.get("category") or "Unknown",
                "evidence_type": evidence["evidence_type"],
                "extracted_evidence": evidence["snippet"],
                "evidence_quality": evidence["quality"],
                "extraction_status": evidence["extraction_status"],
                "relative_path": path,
            }
        )
    return rows


def build_deterministic_ticker_memo(
    ticker: str,
    files: pd.DataFrame,
    snippets: List[Dict[str, Any]],
    *,
    report_date: str,
    mode: str,
) -> str:
    """Build a source-grounded ticker memo without inferring financial conclusions."""
    evidence_rows = build_ticker_memo_evidence_rows(files, snippets)
    paths = files["relative_path"].astype(str).tolist() if not files.empty else []
    cited_paths = ", ".join(f"`{path}`" for path in paths)
    useful = [row for row in evidence_rows if row["extracted_evidence"]]
    lines = [
        f"# Ticker Investment Memo - {ticker} - {report_date}",
        "",
        f"Memo mode: {mode}",
        "",
        "## Files Reviewed",
        "",
        "| Source/broker | File name | Document type | Category | Extraction status | Relevance score |",
        "|---|---|---|---|---|---:|",
    ]
    snippets_by_path = {str(item.get("relative_path") or ""): item for item in snippets}
    for _, row in files.iterrows():
        path = str(row.get("relative_path") or "")
        status = str(snippets_by_path.get(path, {}).get("text_extraction_status") or "metadata-only")
        lines.append(
            f"| {_markdown_table_text(row.get('source_or_broker') or 'Company / filing source')} | "
            f"`{_markdown_table_text(row.get('file_name') or path)}` | "
            f"{_markdown_table_text(row.get('document_type') or 'other')} | "
            f"{_markdown_table_text(row.get('category') or 'Unknown')} | {status} | "
            f"{int(row.get('relevance_score') or 0)} |"
        )

    lines.extend(["", "## Executive Summary"])
    if files.empty:
        lines.append("No qualifying same-day files were identified for this ticker.")
    else:
        lines.append(
            f"{len(files)} same-day file(s) were reviewed for **{ticker}** across "
            f"{files['document_type'].nunique()} detected document type(s). "
            f"Limited extracted text was available; review source PDFs before making investment conclusions. "
            f"Sources: {cited_paths}"
        )

    lines.extend(["", "## Document Coverage Overview"])
    if files.empty:
        lines.append("No document coverage available.")
    else:
        for document_type, group in files.groupby("document_type", sort=True):
            group_paths = ", ".join(f"`{path}`" for path in group["relative_path"].astype(str))
            lines.append(f"- **{document_type}**: {len(group)} file(s). Sources: {group_paths}")

    lines.extend(
        [
            "",
            "## Key Extracted Evidence",
            "",
            "| Source/broker | File name | Evidence type | Extracted evidence | Extraction status |",
            "|---|---|---|---|---|",
        ]
    )
    for row in evidence_rows:
        snippet = row["extracted_evidence"] or "No investment-useful snippet found in limited extraction."
        lines.append(
            f"| {_markdown_table_text(row['source_or_broker'])} | `{_markdown_table_text(row['file_name'])}` | "
            f"{row['evidence_type']} | \"{_markdown_table_text(snippet)}\" | "
            f"{_markdown_table_text(row['extraction_status'])} |"
        )

    def add_evidence_reference_section(title: str, selected_rows: List[Dict[str, Any]], empty_text: str) -> None:
        lines.extend(["", f"## {title}"])
        if not selected_rows:
            lines.append(empty_text)
            return
        for row in selected_rows:
            lines.append(
                f"- **{row['evidence_type']}** evidence is available in the Key Extracted Evidence table. "
                f"Source: `{row['relative_path']}`"
            )

    broker_rows = [row for row in useful if row["source_or_broker"] != "Company / filing source"]
    credit_rows = [
        row for row in useful
        if row["document_type"] == "credit report" or row["evidence_type"] == "Credit / Liquidity"
    ]
    earnings_rows = [
        row for row in useful
        if row["document_type"] in {
            "earnings release", "earnings presentation", "earnings supplement", "transcript",
            "prepared remarks", "10-Q/10Q", "8-K/8K", "MD&A",
        } or row["evidence_type"] in {"Earnings / Results", "Guidance / Estimates"}
    ]
    add_evidence_reference_section(
        "Broker / Source Views",
        broker_rows,
        "No investment-useful broker/source excerpt was found in the limited extraction.",
    )
    lines.append("For full broker-by-broker comparison, use the Broker Consensus Comparator above.")
    add_evidence_reference_section(
        "Credit / Balance Sheet Notes",
        credit_rows,
        "No verified credit, liquidity, leverage, or debt note was found in the limited extraction.",
    )
    add_evidence_reference_section(
        "Earnings / Operating Notes",
        earnings_rows,
        "No verified earnings or operating note was found in the limited extraction.",
    )

    lines.extend(
        [
            "",
            "## Potential Bullish Evidence",
            "No bullish investment evidence was classified automatically. Review the Key Extracted Evidence table "
            f"and source PDFs before assessing upside. Sources reviewed: {cited_paths}",
            "",
            "## Potential Bearish / Risk Evidence",
            "No bearish or risk evidence was classified automatically. Review the Key Extracted Evidence table "
            f"and source PDFs before assessing downside or risk. Sources reviewed: {cited_paths}",
            "",
            "## Open Questions for Geoff/Mitko",
            "- Do the source PDFs contain explicit rating or price-target changes?",
            "- Do company materials verify any EPS, revenue, margin, guidance, or estimate change?",
            "- Do credit documents identify liquidity, leverage, debt, or covenant concerns?",
            "- Which extracted evidence warrants full-document review?",
            "",
            "## Recommended Next Steps",
            "- Open the highest-relevance source documents and verify all financial figures and conclusions.",
            "- Compare company-provided materials with broker and credit-source commentary.",
            "- Record any verified rating, estimate, guidance, valuation, or credit changes manually.",
            "",
            "## Source References",
        ]
    )
    for row in evidence_rows:
        lines.append(
            f"- `{row['relative_path']}` ({row['document_type']}; {row['source_or_broker']}; "
            f"{row['extraction_status']})"
        )
    return "\n".join(lines)


def build_ticker_memo_evidence_payload(
    ticker: str,
    files: pd.DataFrame,
    snippets: List[Dict[str, Any]],
    *,
    source_date: str = "",
    max_total_chars: int = 40000,
) -> str:
    records = []
    used = 0
    for row in build_ticker_memo_evidence_rows(files, snippets):
        record = dict(row)
        record["verified_text_available"] = bool(row["extracted_evidence"])
        record["limited_text"] = record.pop("extracted_evidence") or INSUFFICIENT_TEXT_NOTICE
        encoded = json.dumps(record, ensure_ascii=False)
        if used + len(encoded) > max_total_chars:
            break
        records.append(record)
        used += len(encoded)
    return json.dumps(
        {
            "ticker": ticker,
            "source_date": source_date,
            "document_coverage": (
                files.groupby(["document_type", "category"], dropna=False).size()
                .reset_index(name="file_count").to_dict(orient="records")
                if not files.empty else []
            ),
            "brokers_or_sources": sorted({
                str(value) for value in files.get("source_or_broker", []) if str(value)
            }),
            "grounding_rule": (
                "Metadata supports file/source/category/document-type observations only. Every financial or "
                "investment claim must be a direct quote from limited_text and cite relative_path."
            ),
            "files": records,
        },
        ensure_ascii=False,
    )


def validate_ticker_memo_grounding(text: str, snippets: List[Dict[str, Any]]) -> bool:
    required = [
        "## Files Reviewed", "## Executive Summary", "## Document Coverage Overview",
        "## Key Extracted Evidence", "## Broker / Source Views", "## Credit / Balance Sheet Notes",
        "## Earnings / Operating Notes", "## Potential Bullish Evidence",
        "## Potential Bearish / Risk Evidence", "## Open Questions for Geoff/Mitko",
        "## Recommended Next Steps", "## Source References",
    ]
    if not all(section in text for section in required):
        return False
    before_evidence, after_evidence = text.split("## Key Extracted Evidence", 1)
    key_section, after_evidence = after_evidence.split("## Broker / Source Views", 1)
    for line in key_section.splitlines():
        stripped = line.strip()
        if not stripped.startswith("|") or "source/broker" in stripped.lower() or bool(re.fullmatch(r"[\s|:-]+", stripped)):
            continue
        cited = [
            item for item in snippets
            if str(item.get("relative_path") or "") in line or str(item.get("file_name") or "") in line
        ]
        if not cited:
            return False
        for quote in [part.strip() for part in re.findall(r'"([^"]+)"', line) if part.strip()]:
            if quote == "No investment-useful snippet found in limited extraction.":
                continue
            if not any(quote.lower() in _normalized_extracted_text(item).lower() for item in cited):
                return False
    if re.search(
        r"\b(?:we|investors?|geoff|mitko|you)\s+should\s+(?:buy|sell)\b"
        r"|\bwe recommend (?:buying|selling)\b",
        text,
        flags=re.IGNORECASE,
    ):
        return False
    if re.search(
        r"(?:^|\n)\s*[-*]?\s*(?:buy|sell)\b|\b(?:recommendation|stance)\s*:\s*(?:buy|sell)\b",
        before_evidence + after_evidence,
        flags=re.IGNORECASE,
    ):
        return False
    if re.search(r'"[^"]{20,}"', before_evidence + after_evidence):
        return False
    source_names = {
        value
        for item in snippets
        for value in [str(item.get("relative_path") or ""), str(item.get("file_name") or "")]
        if value
    }
    claim_text = text.split("## Open Questions for Geoff/Mitko", 1)[0]
    financial_claim_text = "\n".join(
        line for line in claim_text.splitlines()
        if "evidence is available in Key Extracted Evidence" not in line
    )
    if not validate_llm_brief_grounding(financial_claim_text, snippets):
        return False
    for line in claim_text.splitlines():
        stripped = line.strip()
        if (
            not stripped
            or stripped.startswith("#")
            or bool(re.fullmatch(r"[\s|:-]+", stripped))
            or (stripped.startswith("|") and "file" in stripped.lower())
            or stripped.startswith("Memo mode:")
            or stripped == "Limited extracted text was available; review source PDFs before making investment conclusions."
            or stripped == "For full broker-by-broker comparison, use the Broker Consensus Comparator above."
        ):
            continue
        if not any(source in stripped for source in source_names):
            return False
        quotes = [part.strip() for part in re.findall(r'"([^"]+)"', stripped) if part.strip()]
        for quote in quotes:
            if quote == "No investment-useful snippet found in limited extraction.":
                continue
            cited = [
                item for item in snippets
                if str(item.get("relative_path") or "") in stripped or str(item.get("file_name") or "") in stripped
            ]
            if not any(quote.lower() in _normalized_extracted_text(item).lower() for item in cited):
                return False
    return True


def build_cross_day_change_inventory(
    previous: pd.DataFrame,
    current: pd.DataFrame,
) -> pd.DataFrame:
    """Compare two scored inventories using filename identity and metadata changes."""
    if previous.empty and current.empty:
        return pd.DataFrame(columns=CROSS_DAY_CHANGE_COLUMNS)

    def keyed_rows(frame: pd.DataFrame) -> Dict[str, Dict[str, Any]]:
        keyed: Dict[str, Dict[str, Any]] = {}
        if frame.empty:
            return keyed
        for _, row in frame.sort_values(["file_name", "relative_path"]).iterrows():
            key = str(row.get("file_name") or "").strip().lower()
            if key and key not in keyed:
                keyed[key] = row.to_dict()
        return keyed

    previous_rows = keyed_rows(previous)
    current_rows = keyed_rows(current)
    changes: List[Dict[str, Any]] = []

    def append_change(change_type: str, row: Dict[str, Any], previous_score: Any, current_score: Any, reason: str) -> None:
        changes.append(
            {
                "change_type": change_type,
                "ticker": row.get("ticker") or "",
                "category": row.get("category") or "Root",
                "source_or_broker": row.get("source_or_broker") or "",
                "document_type": row.get("document_type") or "other",
                "file_name": row.get("file_name") or "",
                "previous_score": previous_score if previous_score != "" else "",
                "current_score": current_score if current_score != "" else "",
                "priority_level": row.get("priority_level") or "",
                "reason_for_change": reason,
            }
        )

    for key in sorted(current_rows.keys() - previous_rows.keys()):
        row = current_rows[key]
        append_change(
            "new",
            row,
            "",
            int(row.get("relevance_score") or 0),
            f"New {row.get('document_type') or 'document'} in {row.get('category') or 'Root'}.",
        )
    for key in sorted(previous_rows.keys() - current_rows.keys()):
        row = previous_rows[key]
        append_change(
            "removed",
            row,
            int(row.get("relevance_score") or 0),
            "",
            "File was present in the previous-day archive but not the current-day archive.",
        )
    for key in sorted(previous_rows.keys() & current_rows.keys()):
        old = previous_rows[key]
        new = current_rows[key]
        size_changed = float(old.get("file_size_mb") or 0) != float(new.get("file_size_mb") or 0)
        date_changed = str(old.get("modified_date") or "") != str(new.get("modified_date") or "")
        if size_changed or date_changed:
            details = []
            if size_changed:
                details.append("file size changed")
            if date_changed:
                details.append("modified date changed")
            append_change(
                "modified",
                new,
                int(old.get("relevance_score") or 0),
                int(new.get("relevance_score") or 0),
                "; ".join(details).capitalize() + ".",
            )

    previous_expanded = _expand_rows_by_detected_ticker(previous)
    current_expanded = _expand_rows_by_detected_ticker(current)
    previous_high = set(
        previous_expanded.loc[previous_expanded["priority_level"] == "High", "ticker"].astype(str)
    ) if not previous_expanded.empty else set()
    current_high_rows = current_expanded[
        (current_expanded["priority_level"] == "High")
        & (~current_expanded["ticker"].astype(str).isin(previous_high))
    ] if not current_expanded.empty else current_expanded
    for ticker, group in current_high_rows.groupby("ticker") if not current_high_rows.empty else []:
        row = group.sort_values("relevance_score", ascending=False).iloc[0].to_dict()
        append_change(
            "moved_to_high_priority",
            row,
            "",
            int(row.get("relevance_score") or 0),
            f"{ticker} has current high-priority research and had none in the previous-day archive.",
        )

    return pd.DataFrame(changes, columns=CROSS_DAY_CHANGE_COLUMNS)


def build_cross_day_ticker_summary(
    previous: pd.DataFrame,
    current: pd.DataFrame,
    changes: pd.DataFrame,
) -> pd.DataFrame:
    previous_expanded = _expand_rows_by_detected_ticker(previous)
    current_expanded = _expand_rows_by_detected_ticker(current)
    tickers = sorted(
        set(previous_expanded.get("ticker", pd.Series(dtype=str)).astype(str))
        | set(current_expanded.get("ticker", pd.Series(dtype=str)).astype(str))
    )
    tickers = [ticker for ticker in tickers if ticker]
    rows = []
    for ticker in tickers:
        old = previous_expanded[previous_expanded["ticker"].astype(str) == ticker]
        new = current_expanded[current_expanded["ticker"].astype(str) == ticker]
        ticker_changes = changes[changes["ticker"].astype(str) == ticker] if not changes.empty else changes
        new_changes = ticker_changes[ticker_changes["change_type"] == "new"] if not ticker_changes.empty else ticker_changes
        removed_changes = ticker_changes[ticker_changes["change_type"] == "removed"] if not ticker_changes.empty else ticker_changes
        old_sources = {str(x) for x in old.get("source_or_broker", []) if str(x)}
        new_sources = {str(x) for x in new.get("source_or_broker", []) if str(x)}
        old_types = {str(x) for x in old.get("document_type", []) if str(x)}
        new_types = {str(x) for x in new.get("document_type", []) if str(x)}
        current_types = new.get("document_type", pd.Series(dtype=str)).astype(str)
        attention = []
        if len(new) > len(old):
            attention.append(f"file count increased by {len(new) - len(old)}")
        if new_sources - old_sources:
            attention.append("new broker/source coverage")
        if (current_types == "credit report").any() and not (old.get("document_type", pd.Series(dtype=str)).astype(str) == "credit report").any():
            attention.append("new credit research")
        if not old.empty and not new.empty and (new["priority_level"] == "High").any() and not (old["priority_level"] == "High").any():
            attention.append("moved into high priority")
        if old.empty and not new.empty:
            attention.append("new ticker/entity")
        if new.empty and not old.empty:
            attention.append("removed ticker/entity")
        rows.append(
            {
                "ticker": ticker,
                "previous_file_count": len(old),
                "current_file_count": len(new),
                "new_file_count": len(new_changes),
                "removed_file_count": len(removed_changes),
                "new_brokers_or_sources": ", ".join(sorted(new_sources - old_sources)),
                "new_document_types": ", ".join(sorted(new_types - old_types)),
                "new_credit_report_count": int((new_changes["document_type"] == "credit report").sum()) if not new_changes.empty else 0,
                "new_earnings_file_count": int(new_changes["document_type"].astype(str).str.startswith("earnings").sum()) if not new_changes.empty else 0,
                "new_transcript_count": int((new_changes["document_type"] == "transcript").sum()) if not new_changes.empty else 0,
                "new_filing_count": int(new_changes["document_type"].isin(["10-Q/10Q", "8-K/8K", "current report", "MD&A"]).sum()) if not new_changes.empty else 0,
                "current_high_priority_count": int((new.get("priority_level", pd.Series(dtype=str)) == "High").sum()),
                "attention_reason": "; ".join(attention) if attention else "No material metadata change detected.",
            }
        )
    result = pd.DataFrame(rows, columns=CROSS_DAY_TICKER_COLUMNS)
    return result.sort_values(
        ["new_file_count", "current_high_priority_count", "ticker"],
        ascending=[False, False, True],
    ).reset_index(drop=True) if not result.empty else result


def build_deterministic_cross_day_report(
    changes: pd.DataFrame,
    ticker_summary: pd.DataFrame,
    *,
    previous_date: str,
    current_date: str,
) -> str:
    lines = [
        f"# Cross-Day Research Change Report - Previous: {previous_date} vs Current: {current_date}",
        "",
        "Generation method: deterministic_fallback",
        "",
        "## Executive Summary",
        (
            f"Detected {int((changes['change_type'] == 'new').sum()) if not changes.empty else 0} new file(s), "
            f"{int((changes['change_type'] == 'removed').sum()) if not changes.empty else 0} removed file(s), and "
            f"{int((changes['change_type'] == 'modified').sum()) if not changes.empty else 0} modified file(s). "
            "This report is metadata-based and does not make financial conclusions."
        ),
    ]

    def add_change_section(title: str, subset: pd.DataFrame, empty_text: str) -> None:
        lines.extend(["", f"## {title}"])
        if subset.empty:
            lines.append(empty_text)
            return
        for _, row in subset.head(30).iterrows():
            lines.append(
                f"- `{row['file_name']}`: {row['reason_for_change']} "
                f"Ticker: {row['ticker'] or 'uncertain'}; source: {row['source_or_broker'] or 'unknown'}; "
                f"document type: {row['document_type']}."
            )

    add_change_section("Key Changes", changes, "No metadata changes were detected.")
    add_change_section(
        "New High-Priority Documents",
        changes[(changes["change_type"] == "new") & (changes["priority_level"] == "High")] if not changes.empty else changes,
        "No new high-priority documents were detected.",
    )
    lines.extend(["", "## Tickers With Increased Attention"])
    attention = ticker_summary[
        (ticker_summary["current_file_count"] > ticker_summary["previous_file_count"])
        | (ticker_summary["new_brokers_or_sources"].astype(str) != "")
        | ticker_summary["attention_reason"].astype(str).str.contains("high priority", case=False)
    ] if not ticker_summary.empty else ticker_summary
    if attention.empty:
        lines.append("No ticker/entity showed increased metadata attention.")
    else:
        for _, row in attention.head(30).iterrows():
            lines.append(f"- **{row['ticker']}**: {row['attention_reason']}.")
    add_change_section(
        "New Broker/Source Coverage",
        changes[(changes["change_type"] == "new") & (changes["source_or_broker"].astype(str) != "")] if not changes.empty else changes,
        "No new broker/source coverage was detected.",
    )
    add_change_section(
        "New Credit Reports",
        changes[(changes["change_type"] == "new") & (changes["document_type"] == "credit report")] if not changes.empty else changes,
        "No new credit reports were detected.",
    )
    add_change_section(
        "New Earnings/Transcript/Filing Materials",
        changes[
            (changes["change_type"] == "new")
            & (
                changes["document_type"].astype(str).str.startswith("earnings")
                | changes["document_type"].isin(["transcript", "10-Q/10Q", "8-K/8K", "current report", "MD&A", "prepared remarks"])
            )
        ] if not changes.empty else changes,
        "No new earnings, transcript, or filing materials were detected.",
    )
    add_change_section(
        "Green Street/Sector Updates",
        changes[
            (changes["change_type"] == "new")
            & (
                (changes["category"] == "Green Street")
                | changes["document_type"].isin(["Green Street report", "sector report", "recommendation change"])
            )
        ] if not changes.empty else changes,
        "No new Green Street, sector, or recommendation updates were detected.",
    )
    add_change_section(
        "Removed or Missing Items",
        changes[changes["change_type"] == "removed"] if not changes.empty else changes,
        "No removed or missing files were detected.",
    )
    lines.extend(
        [
            "",
            "## Recommended Follow-Up for Geoff/Mitko",
            "- Review new high-priority documents and tickers with increased file or source coverage.",
            "- Verify removed files were intentionally absent from the current-day archive.",
            "- Open source documents before drawing conclusions about ratings, results, guidance, or credit.",
            "",
            "## Source References",
        ]
    )
    if changes.empty:
        lines.append("- No changed source filenames.")
    else:
        for name in changes["file_name"].astype(str).drop_duplicates().head(100):
            lines.append(f"- `{name}`")
    return "\n".join(lines)


def build_cross_day_evidence_payload(
    changes: pd.DataFrame,
    ticker_summary: pd.DataFrame,
    *,
    previous_date: str,
    current_date: str,
    max_changes: int = 80,
    max_tickers: int = 50,
) -> str:
    return json.dumps(
        {
            "previous_date": previous_date,
            "current_date": current_date,
            "grounding_rule": "Metadata changes only. Do not infer financial conclusions.",
            "change_inventory": changes.head(int(max_changes)).to_dict(orient="records"),
            "ticker_change_summary": ticker_summary.head(int(max_tickers)).to_dict(orient="records"),
        },
        ensure_ascii=False,
    )


def validate_cross_day_report_grounding(text: str) -> bool:
    required = [
        "## Executive Summary", "## Key Changes", "## New High-Priority Documents",
        "## Tickers With Increased Attention", "## New Broker/Source Coverage",
        "## New Credit Reports", "## New Earnings/Transcript/Filing Materials",
        "## Green Street/Sector Updates", "## Removed or Missing Items",
        "## Recommended Follow-Up for Geoff/Mitko", "## Source References",
    ]
    if not all(section in text for section in required):
        return False
    claim_text = text.split("## Recommended Follow-Up for Geoff/Mitko", 1)[0]
    financial_patterns = [pattern for pattern in SENSITIVE_FINANCE_PATTERNS if "miss" not in pattern]
    return not any(re.search(pattern, claim_text, flags=re.IGNORECASE) for pattern in financial_patterns)


def build_research_index_rows(
    relevance: pd.DataFrame,
    snippets: List[Dict[str, Any]],
    *,
    source_name: str,
) -> pd.DataFrame:
    """Build one lightweight index row per research file without rescanning files."""
    if not isinstance(relevance, pd.DataFrame) or relevance.empty:
        return pd.DataFrame(columns=RESEARCH_INDEX_COLUMNS)

    snippet_by_path: Dict[str, Dict[str, Any]] = {}
    for item in snippets or []:
        path = str(item.get("relative_path") or "")
        if not path:
            continue
        evidence = best_investment_useful_snippet(item, max_chars=900)
        current = snippet_by_path.get(path)
        current_length = len(str(current.get("snippet") or "")) if current else -1
        if len(str(evidence.get("snippet") or "")) > current_length:
            snippet_by_path[path] = evidence

    source_date = daily_source_title_suffix(source_name, relevance.get("relative_path", []))
    rows = []
    for _, row in relevance.iterrows():
        relative_path = str(row.get("relative_path") or "")
        evidence = snippet_by_path.get(relative_path, {})
        rows.append(
            {
                "source_date": source_date,
                "indexed_source": str(source_name or "daily_research.zip"),
                "ticker": str(row.get("ticker") or ""),
                "all_detected_tickers": str(row.get("all_detected_tickers") or row.get("ticker") or ""),
                "source_or_broker": str(row.get("source_or_broker") or ""),
                "category": str(row.get("category") or "Root"),
                "document_type": str(row.get("document_type") or "other"),
                "file_name": str(row.get("file_name") or ""),
                "relative_path": relative_path,
                "relevance_score": int(row.get("relevance_score") or 0),
                "priority_level": str(row.get("priority_level") or ""),
                "extracted_snippet": str(evidence.get("snippet") or ""),
                "evidence_type": str(evidence.get("evidence_type") or "Other"),
                "extraction_status": str(
                    evidence.get("extraction_status")
                    or row.get("extraction_status")
                    or "metadata-only"
                ),
            }
        )
    return pd.DataFrame(rows, columns=RESEARCH_INDEX_COLUMNS)


def build_cross_day_research_index_rows(
    changes: pd.DataFrame,
    *,
    source_name: str,
) -> pd.DataFrame:
    """Index cross-day change metadata so change-oriented queries remain searchable."""
    if not isinstance(changes, pd.DataFrame) or changes.empty:
        return pd.DataFrame(columns=RESEARCH_INDEX_COLUMNS)
    source_date = daily_source_title_suffix(source_name)
    rows = []
    for _, row in changes.iterrows():
        reason = str(row.get("reason_for_change") or "")
        change_type = str(row.get("change_type") or "change")
        source = str(row.get("source_or_broker") or "")
        if change_type == "new" and source:
            reason = f"New broker/source coverage candidate from {source}. {reason}"
        rows.append(
            {
                "source_date": source_date,
                "indexed_source": f"Cross-day changes: {source_name}",
                "ticker": str(row.get("ticker") or ""),
                "all_detected_tickers": str(row.get("ticker") or ""),
                "source_or_broker": source,
                "category": str(row.get("category") or "Root"),
                "document_type": str(row.get("document_type") or "other"),
                "file_name": str(row.get("file_name") or ""),
                "relative_path": str(row.get("file_name") or ""),
                "relevance_score": int(row.get("current_score") or row.get("previous_score") or 0),
                "priority_level": str(row.get("priority_level") or ""),
                "extracted_snippet": f"{change_type}: {reason}".strip(),
                "evidence_type": "Metadata Change",
                "extraction_status": "metadata-only",
            }
        )
    return pd.DataFrame(rows, columns=RESEARCH_INDEX_COLUMNS)


def merge_research_index(existing: Any, incoming: pd.DataFrame) -> pd.DataFrame:
    """Merge index rows, preserving useful snippets and avoiding rerun duplicates."""
    frames = []
    if isinstance(existing, pd.DataFrame) and not existing.empty:
        frames.append(existing.reindex(columns=RESEARCH_INDEX_COLUMNS))
    if isinstance(incoming, pd.DataFrame) and not incoming.empty:
        frames.append(incoming.reindex(columns=RESEARCH_INDEX_COLUMNS))
    if not frames:
        return pd.DataFrame(columns=RESEARCH_INDEX_COLUMNS)
    combined = pd.concat(frames, ignore_index=True)
    combined["_snippet_length"] = combined["extracted_snippet"].astype(str).str.len()
    combined = combined.sort_values("_snippet_length").drop_duplicates(
        subset=["indexed_source", "relative_path"], keep="last"
    )
    return combined.drop(columns=["_snippet_length"]).reset_index(drop=True)


def build_indexed_sources_summary(index_df: pd.DataFrame) -> pd.DataFrame:
    columns = ["source_date", "indexed_source", "files_indexed", "snippets_indexed"]
    if not isinstance(index_df, pd.DataFrame) or index_df.empty:
        return pd.DataFrame(columns=columns)
    summary = (
        index_df.assign(_has_snippet=index_df["extracted_snippet"].astype(str).str.strip().ne(""))
        .groupby(["source_date", "indexed_source"], dropna=False)
        .agg(files_indexed=("relative_path", "nunique"), snippets_indexed=("_has_snippet", "sum"))
        .reset_index()
    )
    return summary.sort_values(["source_date", "indexed_source"], ascending=[False, True]).reset_index(drop=True)


def build_dashboard_priority_tickers(relevance: pd.DataFrame) -> pd.DataFrame:
    """Create a deterministic executive ranking from existing scored inventory."""
    if not isinstance(relevance, pd.DataFrame) or relevance.empty:
        return pd.DataFrame(columns=DASHBOARD_PRIORITY_COLUMNS)
    summary = build_ticker_summary(relevance)
    if summary.empty:
        return pd.DataFrame(columns=DASHBOARD_PRIORITY_COLUMNS)
    rows = []
    for _, row in summary.iterrows():
        broker_report_count = int(row.get("broker_report_count") or 0)
        broker_count = len(
            [source for source in str(row.get("sources") or "").split(",") if source.strip()]
        )
        credit_count = int(row.get("credit_report_count") or 0)
        coverage_count = (
            int(row.get("earnings_file_count") or 0)
            + int(row.get("transcript_count") or 0)
            + int(row.get("filing_file_count") or 0)
        )
        high_count = int(row.get("high_priority_files") or 0)
        attention_score = min(
            100,
            int(row.get("max_relevance_score") or 0)
            + min(high_count * 3, 12)
            + min(broker_report_count * 2, 10)
            + min(credit_count * 3, 9)
            + min(coverage_count * 2, 10),
        )
        rows.append(
            {
                "ticker": row.get("ticker") or "",
                "attention_score": attention_score,
                "file_count": int(row.get("file_count") or 0),
                "high_priority_count": high_count,
                "broker_source_count": broker_count,
                "credit_report_count": credit_count,
                "earnings_transcript_filing_count": coverage_count,
                "categories_present": row.get("categories") or "",
                "attention_reason": row.get("attention_reason") or "",
            }
        )
    return pd.DataFrame(rows, columns=DASHBOARD_PRIORITY_COLUMNS).sort_values(
        ["attention_score", "high_priority_count", "file_count", "ticker"],
        ascending=[False, False, False, True],
    ).reset_index(drop=True)


def _dashboard_signal_types(row: pd.Series) -> List[str]:
    text = " ".join(
        str(row.get(column) or "")
        for column in ["extracted_snippet", "document_type", "category", "file_name", "evidence_type"]
    )
    types = [
        signal_type
        for signal_type, patterns in DASHBOARD_SIGNAL_PATTERNS
        if any(re.search(pattern, text, flags=re.IGNORECASE) for pattern in patterns)
    ]
    document_type = str(row.get("document_type") or "")
    category = str(row.get("category") or "")
    source = str(row.get("source_or_broker") or "")
    if source and document_type in {"analyst report", "credit report", "rating change", "recommendation change"}:
        types.append("Broker Coverage")
    if category == "Green Street" or source == "Green Street" or document_type in {
        "Green Street report", "sector report", "recommendation change",
    }:
        types.append("Green Street / Sector")
    if document_type in {"transcript", "10-Q/10Q", "8-K/8K", "current report", "MD&A", "prepared remarks"}:
        types.append("Filing / Transcript")
    return list(dict.fromkeys(types)) or ["Other"]


def _dashboard_signal_direction(signal_type: str, snippet: str) -> Tuple[str, str]:
    if signal_type == "Rating / Price Target" and re.search(
        r"\b(?:maintain(?:s|ed)?|reiterate[sd]?)\b", snippet, flags=re.IGNORECASE
    ) and not re.search(
        r"\b(?:upgrade[sd]?|downgrade[sd]?|raise[sd]?|lowered|cut(?:s)?)\b",
        snippet,
        flags=re.IGNORECASE,
    ):
        return "Neutral", r"\b(?:maintain(?:s|ed)?|reiterate[sd]?)\b"
    rules = DASHBOARD_DIRECTION_PATTERNS.get(signal_type, {})
    positive = [
        pattern for pattern in rules.get("Positive", ())
        if re.search(pattern, snippet, flags=re.IGNORECASE)
    ]
    negative = [
        pattern for pattern in rules.get("Negative", ())
        if re.search(pattern, snippet, flags=re.IGNORECASE)
    ]
    if positive and not negative:
        return "Positive", positive[0]
    if negative and not positive:
        return "Negative", negative[0]
    return "Unknown", ""


def _dashboard_signal_priority(
    signal_type: str,
    direction: str,
    confidence: str,
    evidence: str,
    source_priority: str,
) -> str:
    evidence_lower = evidence.casefold()
    if signal_type == "Broker Coverage":
        return "Low"
    critical_types = {
        "Rating / Price Target", "Guidance / Estimates", "Credit / Liquidity", "Earnings / Results",
    }
    critical_terms = (
        "upgrade", "downgrade", "price target", "pt raised", "pt cut", "guidance",
        "estimate", "liquidity pressure", "rating pressure", "beat", "miss",
    )
    if (
        signal_type in critical_types
        and direction in {"Positive", "Negative"}
        and any(term in evidence_lower for term in critical_terms)
    ):
        return "Critical"
    if (
        (confidence == "High" and direction in {"Positive", "Negative"})
        or (signal_type in {"Credit / Liquidity", "Earnings / Results"} and confidence == "Medium")
        or (
            signal_type in {"Capex / Investment Spend", "Cash Flow / FCF"}
            and direction == "Negative"
        )
    ):
        return "High"
    if confidence == "Medium" or source_priority == "High":
        return "Medium"
    return "Low"


def _dashboard_review_reason(
    signal_type: str,
    direction: str,
    confidence: str,
    priority: str,
    metadata_only: bool,
) -> str:
    if metadata_only:
        return "Metadata-only signal; review the source file because extracted text did not verify details."
    if direction in {"Positive", "Negative"}:
        return f"Verify the explicit {direction.lower()} {signal_type.lower()} language in the source file."
    if direction == "Neutral":
        return "Verify the maintain/reiterate language and its surrounding context."
    if priority in {"Critical", "High"}:
        return "High-priority evidence is present, but the direction or full context requires verification."
    if confidence == "Medium":
        return "Relevant extracted evidence is present, but the direction remains unclear."
    return "Weak or unclear evidence; review only if this topic is currently important."


def build_dashboard_signal_table(index_df: pd.DataFrame) -> pd.DataFrame:
    """Extract conservative signals from already-indexed snippets and metadata."""
    if not isinstance(index_df, pd.DataFrame) or index_df.empty:
        return pd.DataFrame(columns=DASHBOARD_SIGNAL_COLUMNS)
    rows = []
    for _, row in index_df.iterrows():
        snippet = str(row.get("extracted_snippet") or "").strip()
        metadata_only = not bool(snippet)
        company_value = row.get("company_or_identifier")
        if pd.isna(company_value) or not str(company_value).strip():
            company_value = ""
        ticker_values = [
            value.strip()
            for value in str(row.get("all_detected_tickers") or row.get("ticker") or "").split(",")
            if value.strip()
        ] or [str(row.get("ticker") or "Uncertain")]
        signal_types = sorted(
            _dashboard_signal_types(row),
            key=lambda value: DASHBOARD_SIGNAL_TYPE_RANK.get(value, 99),
        )[:2 if snippet else 1]
        for ticker in ticker_values:
            for signal_type in signal_types:
                signal_direction, direction_pattern = (
                    _dashboard_signal_direction(signal_type, snippet) if snippet else ("Unknown", "")
                )
                if snippet and signal_direction != "Unknown":
                    confidence = "High"
                elif snippet:
                    confidence = "Medium"
                else:
                    confidence = "Low"
                evidence = snippet[:900] if snippet else (
                    f"Metadata match only: {row.get('document_type') or 'other'} in "
                    f"{row.get('category') or 'Root'}."
                )
                priority = _dashboard_signal_priority(
                    signal_type,
                    signal_direction,
                    confidence,
                    evidence,
                    str(row.get("priority_level") or ""),
                )
                needs_manual_review = priority in {"Critical", "High", "Medium"}
                why_review = _dashboard_review_reason(
                    signal_type, signal_direction, confidence, priority, metadata_only
                )
                if direction_pattern:
                    why_review += f" Direction rule matched: {direction_pattern}."
                rows.append(
                    {
                        "ticker": ticker,
                        "company_or_identifier": company_value or ticker,
                        "signal_type": signal_type,
                        "signal_direction": signal_direction,
                        "confidence": confidence,
                        "priority": priority,
                        "evidence": evidence,
                        "source_file": row.get("file_name") or row.get("relative_path") or "",
                        "broker_or_source": row.get("source_or_broker") or "",
                        "source_date": row.get("source_date") or "",
                        "category": row.get("category") or "",
                        "document_type": row.get("document_type") or "",
                        "why_it_matters": DASHBOARD_WHY_IT_MATTERS.get(
                            signal_type, DASHBOARD_WHY_IT_MATTERS["Other"]
                        ),
                        "why_review": why_review,
                        "needs_manual_review": needs_manual_review,
                    }
                )
    result = pd.DataFrame(rows, columns=DASHBOARD_SIGNAL_COLUMNS).drop_duplicates(
        subset=["ticker", "signal_type", "source_file", "evidence"]
    )
    result["_type_rank"] = result["signal_type"].map(DASHBOARD_SIGNAL_TYPE_RANK).fillna(99)
    result["_confidence_rank"] = result["confidence"].map({"High": 0, "Medium": 1, "Low": 2}).fillna(3)
    result["_priority_rank"] = result["priority"].map({"Critical": 0, "High": 1, "Medium": 2, "Low": 3}).fillna(4)
    result = (
        result.sort_values(["_priority_rank", "_confidence_rank", "_type_rank"])
        .groupby(["ticker", "source_file", "evidence"], dropna=False, sort=False)
        .head(2)
        .drop(columns=["_type_rank", "_confidence_rank", "_priority_rank"])
    )
    confidence_order = pd.CategoricalDtype(["High", "Medium", "Low"], ordered=True)
    result["confidence"] = result["confidence"].astype(confidence_order)
    return result.sort_values(
        ["priority", "confidence", "ticker", "signal_type"],
        key=lambda column: (
            column.map({"Critical": 0, "High": 1, "Medium": 2, "Low": 3})
            if column.name == "priority" else column
        ),
        ascending=[True, True, True, True],
    ).reset_index(drop=True)


def build_dashboard_manual_review_queue(signals: pd.DataFrame) -> pd.DataFrame:
    """Return the focused review subset from canonical Dashboard signal rows."""
    if not isinstance(signals, pd.DataFrame) or signals.empty:
        return pd.DataFrame(columns=DASHBOARD_MANUAL_REVIEW_COLUMNS)
    review_mask = (
        signals.get("needs_manual_review", pd.Series(False, index=signals.index)).fillna(False).astype(bool)
        | (signals.get("priority", pd.Series("", index=signals.index)).astype(str) == "Low")
    )
    result = signals[review_mask].reindex(columns=DASHBOARD_MANUAL_REVIEW_COLUMNS).copy()
    if result.empty:
        return pd.DataFrame(columns=DASHBOARD_MANUAL_REVIEW_COLUMNS)
    priority_rank = result["priority"].map({"Critical": 0, "High": 1, "Medium": 2, "Low": 3}).fillna(4)
    result = result.assign(_priority_rank=priority_rank)
    return result.sort_values(
        ["_priority_rank", "ticker", "signal_type", "source_file"],
        ascending=[True, True, True, True],
    ).drop(columns=["_priority_rank"]).drop_duplicates(
        subset=["ticker", "signal_type", "source_file", "evidence"], keep="first"
    ).reset_index(drop=True)


def build_dashboard_summary_markdown(
    relevance: pd.DataFrame,
    priority_tickers: pd.DataFrame,
    signals: pd.DataFrame,
    broker_coverage: pd.DataFrame,
    *,
    new_file_count: int = 0,
    cross_day_available: bool = False,
) -> str:
    total_files = len(relevance) if isinstance(relevance, pd.DataFrame) else 0
    high_files = int((relevance.get("priority_level", pd.Series(dtype=str)) == "High").sum()) if total_files else 0
    manual_review = build_dashboard_manual_review_queue(signals)
    review_counts = manual_review["priority"].value_counts() if not manual_review.empty else pd.Series(dtype=int)
    signal_priority_counts = signals["priority"].value_counts() if not signals.empty else pd.Series(dtype=int)
    explicit_signals = int(
        (
            (signals.get("confidence", pd.Series(dtype=str)).astype(str) == "High")
            & signals.get("signal_direction", pd.Series(dtype=str)).astype(str).isin(["Positive", "Negative"])
        ).sum()
    ) if isinstance(signals, pd.DataFrame) and not signals.empty else 0
    metadata_low_count = int(
        (
            signals.get("evidence", pd.Series(dtype=str))
            .astype(str)
            .str.contains("metadata match only", case=False, na=False)
        ).sum()
    ) if isinstance(signals, pd.DataFrame) and not signals.empty else 0
    top_signal_types = (
        signals["signal_type"].astype(str).value_counts().head(3)
        if isinstance(signals, pd.DataFrame) and not signals.empty else pd.Series(dtype=int)
    )
    most_covered = broker_coverage.iloc[0] if isinstance(broker_coverage, pd.DataFrame) and not broker_coverage.empty else None
    lines = [
        "# Investment Research Dashboard",
        "",
        "Daily research priorities, broker coverage, investment signals, and manual review queue.",
        "",
        "## Overview",
        f"- Total files processed: {total_files}",
        f"- High-priority files: {high_files}",
        f"- Tickers/entities covered: {len(priority_tickers)}",
        f"- Signals identified: {len(signals)}",
        f"- High-confidence explicit signals: {explicit_signals}",
        f"- Critical manual review items: {int(review_counts.get('Critical', 0))}",
        f"- High manual review items: {int(review_counts.get('High', 0))}",
        f"- Medium manual review items: {int(review_counts.get('Medium', 0))}",
        f"- Low-priority download-only review items: {int(review_counts.get('Low', 0))}",
        f"- Metadata-only low-priority signals excluded from the main review queue: {metadata_low_count}",
        f"- New cross-day files: {int(new_file_count)}",
        "",
        "## Executive Snapshot",
        (
            "- Top priority tickers: "
            + (
                ", ".join(priority_tickers["ticker"].astype(str).head(3))
                if not priority_tickers.empty else "No ticker ranking available"
            )
        ),
        (
            "- Top signal types: "
            + (
                ", ".join(f"{name} ({int(count)})" for name, count in top_signal_types.items())
                if not top_signal_types.empty else "No signals detected"
            )
        ),
        (
            f"- Most covered ticker: {most_covered['ticker']} "
            f"({int(most_covered['broker_report_count'])} broker/source report(s))"
            if most_covered is not None else "- Most covered ticker: No broker/source coverage available"
        ),
        (
            f"- Cross-day comparison: available ({int(new_file_count)} new file(s))"
            if cross_day_available else "- Cross-day comparison: not available"
        ),
        "",
        "## Top Priority Tickers",
    ]
    if priority_tickers.empty:
        lines.append("- No ticker/entity summary is available.")
    else:
        for _, row in priority_tickers.head(15).iterrows():
            lines.append(
                f"- **{row['ticker']}**: attention score {row['attention_score']}; "
                f"{row['attention_reason']}."
            )
    lines.extend(["", "## Broker Coverage Leaders"])
    if broker_coverage.empty:
        lines.append("- No normalized broker/source coverage is available.")
    else:
        for _, row in broker_coverage.head(15).iterrows():
            lines.append(
                f"- **{row['ticker']}**: {row['broker_report_count']} report(s) from "
                f"{row['brokers_or_sources']}."
            )
    lines.extend(
        [
            "",
            "## Signal Review",
            (
                f"- {explicit_signals} high-confidence explicit-direction signal(s) were found."
            ),
            "- The main manual-review view prioritizes Critical, High, and Medium items.",
            (
                "- Signal priority distribution: "
                f"Critical {int(signal_priority_counts.get('Critical', 0))}, "
                f"High {int(signal_priority_counts.get('High', 0))}, "
                f"Medium {int(signal_priority_counts.get('Medium', 0))}, "
                f"Low {int(signal_priority_counts.get('Low', 0))}."
            ),
            f"- {metadata_low_count} metadata-only low-priority signal(s) remain available in downloads.",
            "- No financial conclusion should be used without checking the cited source file.",
        ]
    )
    return "\n".join(lines)


def search_research_index(
    index_df: pd.DataFrame,
    query: str,
    *,
    ticker: str = "",
    source_or_broker: str = "",
    category: str = "",
    document_type: str = "",
    source_date: str = "",
    evidence_strength: str = "",
    max_results: int = 100,
) -> pd.DataFrame:
    """Rank lightweight index rows using deterministic metadata and keyword matching."""
    if not isinstance(index_df, pd.DataFrame) or index_df.empty:
        return pd.DataFrame(columns=RESEARCH_SEARCH_RESULT_COLUMNS)
    result = index_df.reindex(columns=RESEARCH_INDEX_COLUMNS).copy()

    filters = {
        "source_or_broker": source_or_broker,
        "category": category,
        "document_type": document_type,
    }
    for column, value in filters.items():
        if value:
            result = result[result[column].astype(str).str.casefold() == str(value).casefold()]
    if source_date:
        source_date_casefold = str(source_date).casefold()
        result = result[
            (result["source_date"].astype(str).str.casefold() == source_date_casefold)
            | (result["indexed_source"].astype(str).str.casefold() == source_date_casefold)
        ]
    if ticker:
        ticker_upper = ticker.upper()
        result = result[
            result.apply(
                lambda row: ticker_upper in {
                    value.strip().upper()
                    for value in str(
                        row.get("all_detected_tickers") or row.get("ticker") or ""
                    ).split(",")
                    if value.strip()
                },
                axis=1,
            )
        ]
    if result.empty:
        return pd.DataFrame(columns=RESEARCH_SEARCH_RESULT_COLUMNS)

    query_text = str(query or "").strip().casefold()
    generic_query_terms = {
        "what", "which", "did", "does", "do", "say", "said", "show", "find",
        "about", "mention", "mentions", "mentioned", "report", "reports",
        "broker", "brokers", "research", "source", "sources", "ticker", "tickers",
        "all", "have", "has", "had",
    }
    terms = [
        term for term in re.findall(r"[a-z0-9][a-z0-9&._-]*", query_text)
        if (len(term) > 1 or term.isdigit()) and term not in generic_query_terms
    ]
    indexed_tickers = {
        value.strip().upper()
        for detected in result["all_detected_tickers"].astype(str)
        for value in detected.split(",")
        if value.strip()
    }
    indexed_tickers.update(
        str(value).strip().upper()
        for value in result["ticker"]
        if str(value).strip()
    )
    query_tickers = {term.upper() for term in terms if term.upper() in indexed_tickers}
    broad_ticker_query = bool(
        re.search(r"\b(?:all|which)\s+tickers?\b|\bacross\s+(?:all\s+)?tickers?\b", query_text)
    )
    if query_tickers and not broad_ticker_query:
        result = result[
            result.apply(
                lambda row: bool(
                    query_tickers
                    & {
                        value.strip().upper()
                        for value in str(
                            row.get("all_detected_tickers") or row.get("ticker") or ""
                        ).split(",")
                        if value.strip()
                    }
                ),
                axis=1,
            )
        ]
    if result.empty:
        return pd.DataFrame(columns=RESEARCH_SEARCH_RESULT_COLUMNS)

    indexed_sources = {
        str(value).strip()
        for value in result["source_or_broker"]
        if str(value).strip()
    }
    query_brokers = {
        source for source in indexed_sources
        if re.search(rf"\b{re.escape(source)}\b", str(query or ""), flags=re.IGNORECASE)
    }
    normalized_query_source = detect_source(str(query or ""))
    if normalized_query_source in indexed_sources:
        query_brokers.add(normalized_query_source)

    parsed_dates = {}
    for value in result["source_date"].astype(str).drop_duplicates():
        try:
            parsed_dates[value] = datetime.strptime(value, "%B %d, %Y").toordinal()
        except ValueError:
            continue
    newest_ordinal = max(parsed_dates.values(), default=0)
    oldest_ordinal = min(parsed_dates.values(), default=0)
    scored = []
    for _, row in result.iterrows():
        searchable_fields = {
            "ticker": str(row.get("all_detected_tickers") or row.get("ticker") or "").casefold(),
            "broker": str(row.get("source_or_broker") or "").casefold(),
            "category": str(row.get("category") or "").casefold(),
            "document": str(row.get("document_type") or "").casefold(),
            "filename": str(row.get("file_name") or "").casefold(),
            "path": str(row.get("relative_path") or "").casefold(),
            "snippet": str(row.get("extracted_snippet") or "").casefold(),
            "evidence": str(row.get("evidence_type") or "").casefold(),
        }
        full_text = " ".join(searchable_fields.values())
        relevance_score = int(row.get("relevance_score") or 0)
        score = min(relevance_score / 5.0, 20.0)
        reasons = [f"relevance score {relevance_score}"]
        matched_terms = []
        snippet_terms = []
        metadata_terms = []
        meaningful_phrase = " ".join(terms)
        quoted_phrases = [
            phrase.casefold().strip()
            for phrase in re.findall(r'"([^"]+)"', str(query or ""))
            if phrase.strip()
        ]
        phrase_match = bool(
            (meaningful_phrase and len(terms) >= 2 and meaningful_phrase in full_text)
            or any(phrase in full_text for phrase in quoted_phrases)
        )
        if phrase_match:
            score += 35
            reasons.append("exact phrase match")
        for term in terms:
            matching_fields = [name for name, value in searchable_fields.items() if term in value]
            if matching_fields:
                matched_terms.append(term)
                if "snippet" in matching_fields:
                    snippet_terms.append(term)
                    score += 12
                else:
                    metadata_terms.append(term)
                    score += 5
                if "filename" in matching_fields:
                    score += 6
                if "broker" in matching_fields:
                    score += 6
        detected = {
            value.strip().upper()
            for value in str(row.get("all_detected_tickers") or row.get("ticker") or "").split(",")
            if value.strip()
        }
        exact_tickers = detected & query_tickers
        if exact_tickers:
            score += 90
            reasons.append(f"exact ticker: {', '.join(sorted(exact_tickers))}")
        row_source = str(row.get("source_or_broker") or "").strip()
        exact_broker = bool(row_source and row_source in query_brokers)
        if exact_broker:
            score += 55
            reasons.append(f"exact broker/source: {row_source}")
        if str(row.get("priority_level") or "") == "High":
            score += 15
            reasons.append("high priority")
        cleaned_snippet = re.sub(
            r"\s+",
            " ",
            str(row.get("extracted_snippet") or "").replace("\x00", " "),
        ).strip()
        has_snippet = bool(cleaned_snippet)
        if has_snippet:
            score += 28
            reasons.append("snippet available")
        else:
            score -= 18
            reasons.append("metadata-only match")
        if str(row.get("evidence_type") or "") not in {"", "Other"}:
            score += 5
            reasons.append(str(row.get("evidence_type")))
        row_ordinal = parsed_dates.get(str(row.get("source_date") or ""), 0)
        if row_ordinal and newest_ordinal:
            date_span = max(newest_ordinal - oldest_ordinal, 1)
            recency_boost = 1 + (3 * (row_ordinal - oldest_ordinal) / date_span)
            score += recency_boost
            reasons.append("source-date recency")
        if matched_terms:
            reasons.append(f"keywords: {', '.join(sorted(set(matched_terms)))}")
        if terms and not matched_terms and not exact_tickers and query_text not in full_text:
            continue

        if exact_tickers:
            match_type = "ticker_match"
        elif exact_broker:
            match_type = "broker_match"
        elif phrase_match:
            match_type = "phrase_match"
        elif snippet_terms:
            match_type = "snippet_match"
        else:
            match_type = "metadata_match"
        if has_snippet and (exact_tickers or exact_broker):
            strength = "High"
        elif has_snippet:
            strength = "Medium"
        else:
            strength = "Low"
        if evidence_strength and strength.casefold() != evidence_strength.casefold():
            continue

        output = row.to_dict()
        output["search_score"] = round(score, 1)
        output["match_type"] = match_type
        output["evidence_strength"] = strength
        output["why_matched"] = "; ".join(reasons)
        output["needs_manual_review"] = strength != "High"
        output["matched_keywords"] = ", ".join(sorted(set(matched_terms)))
        output["display_snippet"] = (
            cleaned_snippet[:520] + ("..." if len(cleaned_snippet) > 520 else "")
            if cleaned_snippet else "Metadata-only match; no extracted snippet is available."
        )
        output["match_reason"] = output["why_matched"]
        scored.append(output)
    if not scored:
        return pd.DataFrame(columns=RESEARCH_SEARCH_RESULT_COLUMNS)
    return (
        pd.DataFrame(scored, columns=RESEARCH_SEARCH_RESULT_COLUMNS)
        .sort_values(["search_score", "relevance_score", "file_name"], ascending=[False, False, True])
        .head(int(max_results))
        .reset_index(drop=True)
    )


def validate_research_query_intent(
    query: str,
    index_df: pd.DataFrame,
    *,
    has_active_filters: bool = False,
) -> Tuple[bool, str]:
    """Allow only questions that plausibly relate to the indexed investment research."""
    query_text = str(query or "").strip()
    if not query_text:
        return (True, "") if has_active_filters else (False, "empty_query")

    normalized = query_text.casefold()
    irrelevant_patterns = [
        r"\bhow are you\b", r"\btell me (?:a|another) joke\b", r"\bweather\b",
        r"\bsports?\b", r"\bpolitics?\b", r"\bpresident\b", r"\belection\b",
        r"\btrivia\b", r"\bwho are you\b", r"\bwhat is your name\b",
        r"\bcurrent events?\b", r"\bnews today\b",
    ]
    if any(re.search(pattern, normalized, flags=re.IGNORECASE) for pattern in irrelevant_patterns):
        return False, "irrelevant_query"

    research_patterns = [
        r"\btickers?\b", r"\bentities?\b", r"\bbrokers?\b", r"\bsources?\b",
        r"\bdocuments?\b", r"\bfiles?\b", r"\bresearch\b", r"\breports?\b",
        r"\bcredit\b", r"\bearnings?\b", r"\brevenue\b", r"\beps\b",
        r"\bguidance\b", r"\bestimates?\b", r"\bratings?\b", r"\bprice targets?\b",
        r"\bupgrades?\b", r"\bdowngrades?\b", r"\bliquidity\b", r"\bleverage\b",
        r"\bdebt\b", r"\bcapex\b", r"\bcapital expenditures?\b", r"\bcash flow\b",
        r"\bmargins?\b", r"\bvaluation\b", r"\brisk\b", r"\boutlook\b",
        r"\bfilings?\b", r"\b10[\s-]?q\b", r"\b8[\s-]?k\b", r"\btranscripts?\b",
        r"\bgreen street\b", r"\bsectors?\b", r"\bhigh[\s-]?priority\b",
        r"\bbroker coverage\b", r"\bextracted evidence\b", r"\bsnippets?\b",
    ]
    if any(re.search(pattern, normalized, flags=re.IGNORECASE) for pattern in research_patterns):
        return True, ""
    if has_active_filters:
        return True, ""
    if not isinstance(index_df, pd.DataFrame) or index_df.empty:
        return False, "no_index_match"

    index_terms: set[str] = set()
    for column in [
        "ticker", "all_detected_tickers", "source_or_broker", "category",
        "document_type", "file_name", "relative_path", "evidence_type",
        "extracted_snippet",
    ]:
        if column not in index_df:
            continue
        for value in index_df[column].astype(str):
            cleaned = value.strip().casefold()
            if not cleaned:
                continue
            index_terms.add(cleaned)
            index_terms.update(
                token for token in re.findall(r"[a-z0-9][a-z0-9&._-]*", cleaned)
                if len(token) >= 2
            )

    query_tokens = {
        token for token in re.findall(r"[a-z0-9][a-z0-9&._-]*", normalized)
        if len(token) >= 2
        and token not in {
            "a", "an", "and", "are", "about", "did", "do", "does", "find", "for",
            "how", "in", "is", "me", "mention", "mentions", "of", "on", "say",
            "said", "show", "the", "to", "what", "which", "with", "you",
        }
    }
    if query_tokens & index_terms:
        return True, ""
    if any(term in normalized for term in index_terms if len(term) >= 4):
        return True, ""
    return False, "no_index_match"


def build_deterministic_research_answer(
    query: str,
    results: pd.DataFrame,
    *,
    max_results: int = 10,
) -> str:
    """Build a conservative answer that distinguishes snippets from metadata matches."""
    top = results.head(int(max_results)) if isinstance(results, pd.DataFrame) else pd.DataFrame()
    lines = ["## Answer Summary"]
    if top.empty:
        lines.append("No indexed research rows matched the question and selected filters.")
    else:
        snippet_count = int(top["extracted_snippet"].astype(str).str.strip().ne("").sum())
        metadata_only_count = len(top) - snippet_count
        first_file = str(top.iloc[0].get("file_name") or "")
        strength_counts = top.get("evidence_strength", pd.Series(dtype=str)).value_counts()
        lines.append(
            f"- Metadata shows {len(results)} matching indexed row(s); {snippet_count} top result(s) include "
            f"extracted snippets and {metadata_only_count} top result(s) are metadata-only. [Source: {first_file}]"
        )
        lines.append(
            f"- Evidence confidence among the top results: {int(strength_counts.get('High', 0))} High, "
            f"{int(strength_counts.get('Medium', 0))} Medium, and {int(strength_counts.get('Low', 0))} Low. "
            f"[Source: {first_file}]"
        )
        for _, row in top.head(3).iterrows():
            source = str(row.get("file_name") or "")
            snippet = str(row.get("display_snippet") or row.get("extracted_snippet") or "").strip()
            if snippet:
                excerpt = snippet[:260]
                lines.append(
                    f"- **{row.get('evidence_strength') or 'Low'} confidence**: The extracted snippet from "
                    f"{source} mentions: {excerpt} [Source: {source}]"
                )
            else:
                lines.append(
                    f"- Metadata shows a {row.get('document_type') or 'research'} file from "
                    f"{row.get('source_or_broker') or 'an unidentified source'}. [Source: {source}]"
                )

    lines.extend(["", "## Top Evidence"])
    if top.empty:
        lines.append("No matching evidence.")
    else:
        for _, row in top.head(5).iterrows():
            snippet = str(row.get("display_snippet") or row.get("extracted_snippet") or "").strip()
            if str(row.get("extracted_snippet") or "").strip():
                lines.append(
                    f"- **{row.get('evidence_strength') or 'Medium'} confidence / "
                    f"{row.get('match_type') or 'snippet_match'}**: The extracted snippet from "
                    f"{row['file_name']} mentions: {snippet} "
                    f"Matched keywords: {row.get('matched_keywords') or 'broad research match'}. "
                    f"[Source: {row['file_name']}]"
                )
            else:
                lines.append(
                    f"- **Low confidence / metadata match**: {row['file_name']} is a metadata-only match; "
                    f"{row['document_type']} in {row['category']}, "
                    f"source/broker {row['source_or_broker'] or 'unknown'}. [Source: {row['file_name']}]"
                )

    lines.extend(["", "## Source References"])
    if top.empty:
        lines.append("No matching source files.")
    else:
        for _, row in top.drop_duplicates(["indexed_source", "relative_path"]).iterrows():
            lines.append(f"- Source: {row['file_name']}")

    lines.extend(
        [
            "",
            "## Caveats / Manual Verification",
            "Search ranking is deterministic and uses filenames, metadata, relevance scores, and previously extracted limited snippets.",
            "",
            "A metadata match does not establish a financial claim or broker view. Review source PDFs before relying on ratings, price targets, EPS, revenue, guidance, estimates, margins, liquidity, valuation, or credit conclusions.",
        ]
    )
    return "\n".join(lines)


def validate_openai_rewrite_preservation(
    original: str,
    refined: str,
    *,
    required_headings: Iterable[str],
    allowed_sources: Iterable[str],
) -> Tuple[bool, str]:
    """Reject rewrites that alter the deterministic report's grounded structure or evidence."""
    original_text = str(original or "")
    refined_text = str(refined or "")
    if not refined_text.strip():
        return False, "OpenAI returned an empty rewrite."

    for heading in required_headings:
        if heading not in refined_text:
            return False, f"Required heading was removed: {heading}"
    original_headings = Counter(
        line.strip() for line in original_text.splitlines() if line.strip().startswith("#")
    )
    refined_headings = Counter(
        line.strip() for line in refined_text.splitlines() if line.strip().startswith("#")
    )
    if original_headings != refined_headings:
        return False, "One or more deterministic headings were added, removed, or changed."

    original_table_rows = Counter(
        line.rstrip() for line in original_text.splitlines() if line.lstrip().startswith("|")
    )
    refined_table_rows = Counter(
        line.rstrip() for line in refined_text.splitlines() if line.lstrip().startswith("|")
    )
    if original_table_rows != refined_table_rows:
        return False, "One or more deterministic Markdown table rows were changed."

    citation_pattern = r"\[Source:\s*[^\]\n]+\]"
    original_citations = Counter(re.findall(citation_pattern, original_text, flags=re.IGNORECASE))
    refined_citations = Counter(re.findall(citation_pattern, refined_text, flags=re.IGNORECASE))
    if original_citations != refined_citations:
        return False, "One or more exact source citations were added, removed, or changed."

    sources = {str(source or "").strip() for source in allowed_sources if str(source or "").strip()}
    protected_sources = {source for source in sources if source in original_text}
    if any(refined_text.count(source) != original_text.count(source) for source in protected_sources):
        return False, "One or more source filename occurrences were added, removed, or changed."

    unknown_sources = {
        value.strip()
        for value in re.findall(r"(?:Source:\s*|`)([^`\]\n]+\.(?:pdf|xlsx?|csv|txt|md))", refined_text, flags=re.IGNORECASE)
        if value.strip() not in sources
    }
    if unknown_sources:
        return False, "The rewrite introduced a source filename outside the deterministic report."
    for line in refined_text.splitlines():
        scrubbed = line
        for source in sorted(sources, key=len, reverse=True):
            scrubbed = scrubbed.replace(source, "")
        if re.search(r"\.(?:pdf|xlsx?|csv|txt|md)\b", scrubbed, flags=re.IGNORECASE):
            return False, "The rewrite introduced or changed a filename."

    protected_quotes = Counter(
        quote.strip()
        for quote in re.findall(r'"([^"]+)"', original_text)
        if len(quote.strip()) >= 20
    )
    refined_quotes = Counter(
        quote.strip()
        for quote in re.findall(r'"([^"]+)"', refined_text)
        if len(quote.strip()) >= 20
    )
    if protected_quotes != refined_quotes:
        return False, "Quoted extracted evidence was added, removed, or paraphrased."

    financial_keyword_patterns = {
        "rating": r"\bratings?\b",
        "price target": r"\bprice[-\s]+targets?\b",
        "EPS": r"\beps\b",
        "revenue": r"\brevenue\b",
        "guidance": r"\bguidance\b",
        "estimate": r"\bestimates?\b",
        "capex": r"\bcapex\b|\bcapital[-\s]+expenditures?\b",
        "free cash flow": r"\bfree[-\s]+cash[-\s]+flow\b|\bfcf\b",
        "margin": r"\bmargins?\b",
        "liquidity": r"\bliquidity\b",
        "leverage": r"\bleverage\b",
        "debt": r"\bdebt\b",
        "valuation": r"\bvaluation\b",
        "credit": r"\bcredit\b",
        "upgrade": r"\bupgrades?\b",
        "downgrade": r"\bdowngrades?\b",
        "beat": r"\bbeats?\b",
        "miss": r"\bmiss(?:es|ed|ing)?\b",
        "bullish": r"\bbullish\b",
        "bearish": r"\bbearish\b",
    }
    original_keywords = {
        name for name, pattern in financial_keyword_patterns.items()
        if re.search(pattern, original_text, flags=re.IGNORECASE)
    }
    refined_keywords = {
        name for name, pattern in financial_keyword_patterns.items()
        if re.search(pattern, refined_text, flags=re.IGNORECASE)
    }
    new_keywords = sorted(refined_keywords - original_keywords)
    if new_keywords:
        return False, "The rewrite introduced new financial claim keyword(s): " + ", ".join(new_keywords)
    return True, ""


def validate_openai_rewrite_new_claims(
    original: str,
    refined: str,
    snippets: List[Dict[str, Any]],
) -> Tuple[bool, str]:
    """Apply strict grounding only to lines newly introduced by a constrained rewrite."""
    advice_pattern = (
        r"\b(?:we|investors?|geoff|mitko|you)\s+should\s+(?:buy|sell)\b"
        r"|\bwe recommend (?:buying|selling)\b"
        r"|(?:^|\n)\s*[-*]?\s*(?:buy|sell)\s+(?:recommendation|rating)?\b"
    )
    if re.search(advice_pattern, refined, flags=re.IGNORECASE):
        return False, "The rewrite introduced investment advice."

    original_lines = Counter(line.strip() for line in str(original or "").splitlines() if line.strip())
    refined_lines = Counter(line.strip() for line in str(refined or "").splitlines() if line.strip())
    added_lines = [
        line
        for line, count in refined_lines.items()
        for _ in range(max(0, count - original_lines[line]))
    ]
    if added_lines and not validate_llm_brief_grounding("\n".join(added_lines), snippets):
        return False, "The rewrite introduced an unsupported financial claim."
    return True, ""


def validate_openai_polish_grounding(
    refined: str,
    *,
    context_text: str,
    allowed_sources: Iterable[str],
    allowed_tickers: Iterable[str] = (),
    allowed_brokers: Iterable[str] = (),
) -> Tuple[bool, str]:
    """Apply a practical safety gate to readability-focused OpenAI polish."""
    text = str(refined or "").strip()
    context = str(context_text or "")
    if len(text) < 180 or len([line for line in text.splitlines() if line.strip()]) < 4:
        return False, "OpenAI polish was empty or too short."

    source_heading = re.search(r"(?im)^#{1,4}\s+(?:Sources Used|Source References)\s*$", text)
    if not source_heading:
        return False, "OpenAI polish did not include a Sources Used or Source References section."

    sources = {
        Path(str(source or "").replace("\\", "/")).name.strip()
        for source in allowed_sources
        if str(source or "").strip()
    }
    mentioned_sources = {
        source for source in sources if source and source.casefold() in text.casefold()
    }
    if not mentioned_sources:
        return False, "OpenAI polish did not include a recognizable retrieved source filename."
    source_section_tail = text[source_heading.end():]
    next_heading = re.search(r"(?m)^#{1,4}\s+\S", source_section_tail)
    source_section = source_section_tail[:next_heading.start()] if next_heading else source_section_tail
    if not any(source.casefold() in source_section.casefold() for source in sources):
        return False, "Source References did not include an exact retrieved source filename."

    source_names_casefold = {source.casefold() for source in sources}
    unknown_file_lines = [
        line.strip() for line in text.splitlines()
        if re.search(r"\.(?:pdf|xlsx?|csv|txt|md)\b", line, flags=re.IGNORECASE)
        and not any(source in line.casefold() for source in source_names_casefold)
    ]
    if unknown_file_lines:
        return False, "OpenAI polish introduced source filename(s) outside the retrieved context."

    advice_pattern = (
        r"\b(?:we|investors?|geoff|mitko|you)\s+should\s+(?:buy|sell|hold)\b"
        r"|\bwe recommend (?:buying|selling|holding)\b"
        r"|\b(?:buy|sell|hold)\s+recommendation\b"
    )
    if re.search(advice_pattern, text, flags=re.IGNORECASE):
        return False, "OpenAI polish introduced a buy/sell/hold recommendation."
    if re.search(
        r"\b(?:reviewed|read|analyzed|analysed|examined)\s+(?:the\s+)?(?:full|entire)\s+"
        r"(?:pdfs?|documents?|reports?)\b",
        text,
        flags=re.IGNORECASE,
    ):
        return False, "OpenAI polish incorrectly claimed that full PDFs or documents were reviewed."

    known_tickers = {str(value).strip().upper() for value in allowed_tickers if str(value).strip()}
    ticker_stopwords = TICKER_STOPWORDS | {
        "AI", "API", "CSV", "PDF", "PDFS", "Q&A", "REITS", "SEC", "PT",
        "AWS", "FCF", "MD", "RAG", "LLM", "OPENAI", "DCF", "WACC", "ROIC",
        "ROE", "EBIT", "CAGR", "NIM", "NII", "CRE", "YOY", "QOQ", "YTD",
        "LTM", "NTM", "GAAP",
    }
    ticker_candidates = {
        token.upper()
        for token in re.findall(r"(?<![A-Za-z0-9])\$?([A-Z]{2,5})(?![A-Za-z0-9])", text)
        if token.upper() not in ticker_stopwords
    }
    context_ticker_candidates = {
        token.upper()
        for token in re.findall(r"(?<![A-Za-z0-9])\$?([A-Z]{2,5})(?![A-Za-z0-9])", context)
        if token.upper() not in ticker_stopwords
    }
    fabricated_tickers = sorted(ticker_candidates - known_tickers - context_ticker_candidates)
    if fabricated_tickers:
        return False, "OpenAI polish introduced ticker-like identifier(s) outside the retrieved context."

    known_brokers = {str(value).strip().casefold() for value in allowed_brokers if str(value).strip()}
    context_casefold = context.casefold()
    fabricated_brokers = sorted(
        label for label, pattern in BROKER_PATTERNS
        if re.search(pattern, text, flags=re.IGNORECASE)
        and label.casefold() not in known_brokers
        and not re.search(pattern, context_casefold, flags=re.IGNORECASE)
    )
    if fabricated_brokers:
        return False, "OpenAI polish introduced broker/source name(s) outside the retrieved context."

    month_pattern = (
        r"(?:January|February|March|April|May|June|July|August|September|October|November|December)"
    )

    def date_signatures(value: str) -> set[str]:
        signatures = set()
        for month, day, year in re.findall(
            rf"\b({month_pattern})\s+(\d{{1,2}}),?\s+(\d{{2,4}})\b",
            value,
            flags=re.IGNORECASE,
        ):
            normalized_year = int(year) + (2000 if len(year) == 2 else 0)
            try:
                parsed = datetime.strptime(f"{month} {day} {normalized_year}", "%B %d %Y")
                signatures.add(parsed.strftime("%Y-%m-%d"))
            except ValueError:
                pass
        for month, day, year in re.findall(r"\b(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})\b", value):
            normalized_year = int(year) + (2000 if len(year) == 2 else 0)
            try:
                signatures.add(datetime(normalized_year, int(month), int(day)).strftime("%Y-%m-%d"))
            except ValueError:
                pass
        return signatures

    fabricated_dates = sorted(date_signatures(text) - date_signatures(context))
    if fabricated_dates:
        return False, "OpenAI polish introduced date(s) outside the retrieved context."

    relevance_stopwords = {
        "about", "answer", "based", "business", "caveat", "clear", "concise",
        "context", "document", "documents", "evidence", "files", "information",
        "metadata", "only", "polish", "report", "reports", "research", "retrieved",
        "review", "section", "snippet", "snippets", "source", "sources", "summary",
        "used",
    }

    def relevance_terms(value: str) -> set[str]:
        scrubbed = value
        for source in sources:
            scrubbed = re.sub(re.escape(source), " ", scrubbed, flags=re.IGNORECASE)
        return {
            word.casefold() for word in re.findall(r"\b[A-Za-z][A-Za-z0-9&.-]{4,}\b", scrubbed)
            if word.casefold() not in relevance_stopwords
        }

    if not (relevance_terms(text) & relevance_terms(context)):
        return False, "OpenAI polish was clearly unrelated to the retrieved evidence."

    return True, ""


def build_research_answer_payload(query: str, results: pd.DataFrame, *, max_results: int) -> str:
    fields = [
        "source_date", "ticker", "all_detected_tickers", "source_or_broker", "category",
        "document_type", "file_name", "relative_path", "relevance_score", "priority_level",
        "extracted_snippet", "evidence_type", "extraction_status", "match_type",
        "evidence_strength", "why_matched", "matched_keywords", "display_snippet",
    ]
    compact = results.reindex(columns=fields).head(int(max_results)).copy()
    compact["extracted_snippet"] = compact["extracted_snippet"].astype(str).str.slice(0, 900)
    return json.dumps(
        {"question": query, "results": compact.to_dict(orient="records")},
        ensure_ascii=False,
    )


def validate_research_answer_grounding_detailed(
    text: str,
    results: pd.DataFrame,
) -> Tuple[bool, str, str]:
    """Validate Historical Q&A answers and return a user-readable failure reason."""
    required = [
        "## Answer Summary", "## Top Evidence",
        "## Source References", "## Caveats / Manual Verification",
    ]
    if not str(text or "").strip():
        return False, "empty response", "OpenAI returned no answer text."
    heading_lines = [line.strip() for line in text.splitlines() if line.strip().startswith("#")]
    missing_sections = [section for section in required if section not in heading_lines]
    if missing_sections:
        return (
            False,
            "parsing/format issue",
            "Missing required section(s): " + ", ".join(missing_sections),
        )
    unexpected_headings = [heading for heading in heading_lines if heading not in required]
    if unexpected_headings:
        return (
            False,
            "parsing/format issue",
            "Unexpected heading(s) found. Use only the required `##` headings: "
            + ", ".join(unexpected_headings[:10]),
        )
    if not isinstance(results, pd.DataFrame) or results.empty:
        return False, "parsing/format issue", "No top search results were available for validation."

    rows = results.to_dict(orient="records")
    source_names = {
        str(row.get("file_name") or "").strip()
        for row in rows
        if str(row.get("file_name") or "").strip()
    }
    if not source_names:
        return False, "parsing/format issue", "Top search results did not contain source filenames."

    citation_blocks = re.findall(
        r"(?:\[?Source:\s*)([^\]\n]+?)(?:\]|$)",
        text,
        flags=re.IGNORECASE,
    )
    cited_labels = [
        label.strip().strip("`").strip()
        for block in citation_blocks
        for label in block.split(";")
        if label.strip()
    ]
    unknown_labels = sorted({label for label in cited_labels if label not in source_names})
    if unknown_labels:
        return (
            False,
            "unknown source",
            "Cited source filename(s) were not present in the provided top results: "
            + ", ".join(unknown_labels[:10]),
        )
    mentioned_files = {
        match.strip()
        for match in re.findall(
            r"`([^`\n]+\.(?:pdf|xlsx?|csv|txt|md))`",
            text,
            flags=re.IGNORECASE,
        )
        if match.strip()
    }
    unknown_mentions = sorted(
        mentioned for mentioned in mentioned_files
        if mentioned not in source_names
    )
    if unknown_mentions:
        return (
            False,
            "unknown source",
            "Mentioned source filename(s) were not present in the provided top results: "
            + ", ".join(unknown_mentions[:10]),
        )

    buy_sell_pattern = (
        r"\b(?:we|investors?|geoff|mitko|you)\s+should\s+(?:buy|sell)\b"
        r"|\bwe recommend (?:buying|selling)\b"
        r"|(?:^|\n)\s*[-*]?\s*(?:buy|sell)\s+(?:recommendation|rating)?\b"
    )
    if re.search(buy_sell_pattern, text, flags=re.IGNORECASE):
        return False, "unsupported claim", "The answer included a prohibited buy/sell recommendation."

    research_sensitive_patterns = [
        *SENSITIVE_FINANCE_PATTERNS,
        r"\bratings?\b", r"\bprice targets?\b", r"\bestimates?\b",
        r"\b(?:buy|sell|outperform|underperform|overweight|underweight|market weight|neutral)\b",
        r"\bleverage\b", r"\bdebt\b", r"\bcash flow\b", r"\bcredit concern\b",
    ]
    section = ""
    substantive_sections = {"## Answer Summary", "## Top Evidence"}
    for line_number, line in enumerate(text.splitlines(), start=1):
        stripped = line.strip()
        if stripped in required:
            section = stripped
            continue
        if not stripped or stripped.startswith("#") or stripped.startswith("Generation method:"):
            continue

        cited_rows = [
            row for row in rows
            if str(row.get("file_name") or "") and str(row.get("file_name") or "") in stripped
        ]
        if section in substantive_sections and not cited_rows:
            return (
                False,
                "missing source filenames",
                f"Substantive claim on line {line_number} did not cite an exact top-result filename: {stripped[:240]}",
            )

        sensitive = any(
            re.search(pattern, stripped, flags=re.IGNORECASE)
            for pattern in research_sensitive_patterns
        )
        if not sensitive or section == "## Caveats / Manual Verification":
            continue
        if not cited_rows:
            return (
                False,
                "unsupported claim",
                f"Sensitive financial claim on line {line_number} did not cite a source filename: {stripped[:240]}",
            )
        quotes = [quote.strip() for quote in re.findall(r'"([^"]+)"', stripped) if quote.strip()]
        exact_quote_supported = bool(quotes) and all(
            any(
                quote.casefold() in str(row.get("extracted_snippet") or "").casefold()
                for row in cited_rows
            )
            for quote in quotes
        )
        if exact_quote_supported:
            continue
        cautious_snippet_observation = bool(
            re.search(
                r"\b(?:the\s+)?(?:extracted\s+)?snippets?\s+(?:from\s+.+?\s+)?"
                r"(?:mention|mentions|include|includes|contain|contains)\b",
                stripped,
                flags=re.IGNORECASE,
            )
        )
        concept_patterns = {
            "rating": r"\bratings?\b",
            "price target": r"\bprice[-\s]+targets?\b",
            "EPS": r"\beps\b",
            "revenue": r"\brevenue\b",
            "guidance": r"\bguidance\b",
            "estimate": r"\bestimates?\b",
            "margin": r"\bmargins?\b",
            "operating income": r"\boperating[-\s]+income\b",
            "capex": r"\bcapex\b|\bcapital[-\s]+expenditures?\b",
            "free cash flow": r"\bfree[-\s]+cash[-\s]+flow\b|\bfcf\b",
            "liquidity": r"\bliquidity\b",
            "valuation": r"\bvaluation\b",
            "credit": r"\bcredit\b",
            "cash flow": r"\bcash[-\s]+flow\b",
            "leverage": r"\bleverage\b",
            "debt": r"\bdebt\b",
            "overweight": r"\boverweight\b",
            "underweight": r"\bunderweight\b",
            "upgrade": r"\bupgrades?\b",
            "downgrade": r"\bdowngrades?\b",
            "beat": r"\bbeats?\b",
            "miss": r"\bmiss(?:es|ed)?\b",
        }
        mentioned_concepts = {
            name: pattern for name, pattern in concept_patterns.items()
            if re.search(pattern, stripped, flags=re.IGNORECASE)
        }
        cited_snippets = " ".join(
            str(row.get("extracted_snippet") or "") for row in cited_rows
        )
        concepts_supported = bool(mentioned_concepts) and all(
            re.search(pattern, cited_snippets, flags=re.IGNORECASE)
            for pattern in mentioned_concepts.values()
        )
        prohibited_interpretation = re.search(
            r"\b(?:this suggests|reports? indicate|brokers? (?:noted|believe|conclude)|"
            r"investment case|performance was|appears (?:strong|weak)|is (?:bullish|bearish))\b",
            stripped,
            flags=re.IGNORECASE,
        )
        if cautious_snippet_observation and concepts_supported and not prohibited_interpretation:
            continue
        return (
            False,
            "unsupported claim",
            f"Sensitive financial claim on line {line_number} was neither direct extracted wording nor "
            "a cautious snippet-language observation supported by the cited snippets.",
        )

    evidence_section = text.split("## Top Evidence", 1)[1].split("## Source References", 1)[0]
    summary_section = text.split("## Answer Summary", 1)[1].split("## Top Evidence", 1)[0]
    summary_bullets = [
        line.strip() for line in summary_section.splitlines()
        if line.strip().startswith(("-", "*"))
    ]
    if not 3 <= len(summary_bullets) <= 5:
        return False, "parsing/format issue", "Answer Summary must contain 3-5 concise bullets."
    evidence_bullets = [
        line.strip() for line in evidence_section.splitlines()
        if line.strip().startswith(("-", "*"))
    ]
    if len(evidence_bullets) > 5:
        return False, "parsing/format issue", "Top Evidence must contain no more than 5 bullets."
    if evidence_bullets and not all(any(name in bullet for name in source_names) for bullet in evidence_bullets):
        return False, "missing source filenames", "One or more evidence bullets omitted an exact source filename."
    return True, "", ""


def validate_research_answer_grounding(text: str, results: pd.DataFrame) -> bool:
    valid, _, _ = validate_research_answer_grounding_detailed(text, results)
    return valid


def _attention_reason(group: pd.DataFrame, *, ticker: str) -> str:
    reasons = []
    sources = {str(x) for x in group["source_or_broker"] if str(x)}
    categories = {str(x) for x in group["category"] if str(x)}
    doc_types = group["document_type"].astype(str)
    if len(sources) > 1:
        reasons.append(f"{len(sources)} brokers/sources")
    if len(categories) > 1:
        reasons.append(f"coverage across {len(categories)} categories")
    if (doc_types == "credit report").any():
        reasons.append("credit research present")
    if doc_types.str.startswith("earnings").any() or (doc_types == "transcript").any():
        reasons.append("earnings-related material present")
    if doc_types.isin(["recommendation change", "rating change", "guidance", "estimate revision"]).any():
        reasons.append("change-oriented document present")
    if not reasons:
        reasons.append(f"{len(group)} relevant file(s) for {ticker}")
    return "; ".join(reasons)


def _expand_rows_by_detected_ticker(relevance: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, row in relevance.iterrows():
        tickers_in_row = [
            t.strip() for t in str(row.get("all_detected_tickers") or row.get("ticker") or "").split(",")
            if t.strip()
        ]
        for detected in tickers_in_row:
            expanded = row.to_dict()
            expanded["ticker"] = detected
            rows.append(expanded)
    return pd.DataFrame(rows, columns=relevance.columns) if rows else relevance.iloc[0:0].copy()


def build_category_summary(relevance: pd.DataFrame) -> pd.DataFrame:
    if relevance.empty:
        return pd.DataFrame(columns=["category", "file_count", "high_priority_files", "average_relevance_score", "top_document_types"])
    rows = []
    for category, group in relevance.groupby("category"):
        top_types = group["document_type"].value_counts().head(3).index.tolist()
        rows.append(
            {
                "category": category,
                "file_count": len(group),
                "high_priority_files": int((group["priority_level"] == "High").sum()),
                "average_relevance_score": round(float(group["relevance_score"].mean()), 1),
                "top_document_types": ", ".join(top_types),
            }
        )
    return pd.DataFrame(rows).sort_values(["high_priority_files", "file_count"], ascending=False)


def build_deterministic_brief(
    relevance: pd.DataFrame,
    selected_text: List[Dict[str, Any]],
    *,
    source_name: str,
) -> str:
    ticker_summary = build_ticker_summary(relevance)
    broker_summary = build_broker_coverage_summary(relevance)
    category_summary = build_category_summary(relevance)
    selected_by_path = {str(x.get("relative_path") or ""): x for x in selected_text}
    high = relevance[relevance["priority_level"] == "High"].head(15) if not relevance.empty else relevance
    skipped = relevance[~relevance["relative_path"].isin(selected_by_path.keys())] if not relevance.empty else relevance
    verified_text = verified_extracted_text_items(selected_text)
    lines = [
        f"# Daily Research Brief - {daily_source_title_suffix(source_name, relevance.get('relative_path', []))}",
        "",
        f"Source archive: `{source_name}`",
        f"Files inventoried: {len(relevance)}",
        f"Files lightly scanned: {len(selected_text)}",
        f"High-priority files: {len(relevance[relevance['priority_level'] == 'High']) if not relevance.empty else 0}",
        "",
        "## Metadata-Based Observations",
    ]
    if ticker_summary.empty:
        lines.append("No ticker could be identified conservatively from the filenames.")
    else:
        for _, row in ticker_summary.head(12).iterrows():
            lines.append(
                f"- **{row['ticker']}**: {row['attention_reason']}; "
                f"{row['file_count']} file(s), {row['high_priority_files']} high priority, "
                f"max score {row['max_relevance_score']}."
            )

    lines.extend(["", "### Top High-Priority Documents"])
    if high.empty:
        lines.append("No files reached the deterministic high-priority threshold.")
    else:
        for _, row in high.iterrows():
            lines.append(
                f"- Score {int(row['relevance_score'])}: `{row['relative_path']}` "
                f"({row['document_type']}; {row['source_or_broker'] or 'source unknown'}; "
                f"{row['ticker'] or 'ticker uncertain'}). Rationale: {row['investment_rationale']}"
            )

    lines.extend(["", "### Broker Coverage Highlights"])
    if broker_summary.empty:
        lines.append("No normalized broker/source coverage was identified.")
    else:
        for _, row in broker_summary.head(12).iterrows():
            lines.append(
                f"- **{row['ticker']}**: {row['broker_report_count']} broker/source report(s) from "
                f"{row['brokers_or_sources']}; {row['attention_reason']}."
            )

    lines.extend(["", "### Category Summary"])
    for _, row in category_summary.iterrows():
        lines.append(
            f"- **{row['category']}**: {row['file_count']} file(s), "
            f"{row['high_priority_files']} high priority; common types: {row['top_document_types'] or 'other'}."
        )

    lines.extend(["", "## Extracted-Text Signals"])
    if not verified_text:
        lines.append(INSUFFICIENT_TEXT_NOTICE)
        lines.extend(["", "### Potential Areas to Review"])
        lines.append(
            "Review the high-priority source files listed above for actual results, estimates, "
            "ratings, guidance, and other financial conclusions."
        )
    else:
        lines.append(
            "The following are direct excerpts from successfully extracted source text. "
            "They are not independent conclusions."
        )
        for item in verified_text[:12]:
            text = _normalized_extracted_text(item)[:300]
            excerpt = text + ("..." if len(_normalized_extracted_text(item)) > 300 else "")
            lines.append(f"- `{item['relative_path']}`: extracted-text excerpt: \"{excerpt}\"")

    lines.extend(
        [
            "",
            "## Recommended Follow-Up",
            "For Geoff / Mitko:",
            "- Review the highest-scoring documents first, especially repeated tickers across multiple sources.",
            "- Prioritize tickers with multiple brokers, cross-category coverage, or both credit and equity research.",
            "- Open source files before acting on any possible signal; this brief uses only filenames, metadata, and limited first-page text.",
            "- Validate any ratings, estimates, financial figures, or conclusions directly in the source documents.",
            "",
            "## Skipped or Lightly Scanned Files",
        ]
    )
    if skipped.empty:
        lines.append("All inventoried supported files were included in the lightweight scan.")
    else:
        for _, row in skipped.head(30).iterrows():
            lines.append(f"- `{row['relative_path']}`: inventory/scoring only.")
        if len(skipped) > 30:
            lines.append(f"- ...and {len(skipped) - 30} additional inventory-only file(s).")

    lines.extend(["", "## Source References"])
    reference_paths = set()
    for _, row in high.iterrows():
        path = str(row.get("relative_path") or "")
        if path and path not in reference_paths:
            lines.append(f"- `{path}` (high-priority metadata source)")
            reference_paths.add(path)
    for item in selected_text:
        path = str(item.get("relative_path") or "")
        if path and path not in reference_paths:
            lines.append(f"- `{path}` ({item['text_extraction_status']})")
            reference_paths.add(path)
    return "\n".join(lines)


CUTLER_BRIEF_SECTIONS = [
    "Executive Summary",
    "Macro / Market Backdrop",
    "Convertibles / Credit",
    "REITs / Real Estate",
    "Community Banks / Financials",
    "Healthcare / Pharma",
    "Energy / Industrials / Commodities",
    "Insider Trading / Screens",
    "Other Notable Research",
    "What the Report Should Definitely Catch",
    "Must-Have Items for Review",
    "Actionability / Suggested Follow-Up",
    "Source References",
    "Caveats / Manual Verification",
]


def _cutler_brief_theme(row: Dict[str, Any]) -> str:
    text = " ".join(
        str(row.get(column) or "")
        for column in [
            "ticker", "company_or_identifier", "signal_type", "category", "document_type",
            "source_file", "file_name", "relative_path", "broker_or_source",
            "source_or_broker", "evidence",
        ]
    ).casefold()
    category = str(row.get("category") or "").casefold()
    ticker = str(row.get("ticker") or "").upper()
    if ticker in _CUTLER_TARGETED_REIT_TICKERS:
        return "REITs / Real Estate"
    if ticker in _CUTLER_TARGETED_HEALTHCARE_TICKERS:
        return "Healthcare / Pharma"
    if any(term in text for term in ["reit", "real estate", "green street", "property", "nav"]):
        return "REITs / Real Estate"
    if category == "banks" or any(
        term in text for term in ["community bank", "banking", "bank ", "financials", "financial services"]
    ):
        return "Community Banks / Financials"
    if category == "credit" or any(
        term in text
        for term in [
            "convertible", "convertibles", "credit", "liquidity", "leverage", "debt",
            "covenant", "issuance", "cash flow / fcf",
        ]
    ):
        return "Convertibles / Credit"
    if any(
        term in text
        for term in [
            "healthcare", "pharma", "biotech", "clinical", "fda", "drug", "asco",
            "keytruda", "melanoma", "braftovi",
        ]
    ):
        return "Healthcare / Pharma"
    if any(
        term in text
        for term in [
            "energy", "oil", "gas", "industrial", "commodity", "commodities", "metals",
            "mining", "copper", "refining",
        ]
    ):
        return "Energy / Industrials / Commodities"
    if any(term in text for term in ["insider trading", "screen", "dividend", "income screen"]):
        return "Insider Trading / Screens"
    if any(
        term in text
        for term in [
            "macro", "market backdrop", "fear and greed", "rates", "inflation",
            "jpm guide", "guide to the markets", "market recap", "monday guide",
        ]
    ):
        return "Macro / Market Backdrop"
    return "Other Notable Research"


def _cutler_actionability(row: Dict[str, Any], theme: str, watchlist: set[str]) -> str:
    ticker = str(row.get("ticker") or "").upper()
    signal_type = str(row.get("signal_type") or "")
    evidence = str(row.get("evidence") or "").casefold()
    text = " ".join(
        str(row.get(column) or "")
        for column in ["source_file", "document_type", "category", "signal_type", "evidence"]
    ).casefold()
    if theme == "Community Banks / Financials":
        return "Relevant to community bank strategy"
    if theme == "REITs / Real Estate":
        if signal_type == "Valuation / Thesis" or "nav" in evidence:
            return "Potential NAV estimate impact"
        return "Relevant to REIT holdings/watchlist"
    if theme == "Convertibles / Credit" and any(
        term in text for term in ["convertible", "convertibles", "convert vantage", "dilution"]
    ):
        return "Potential credit/convertible dilution impact"
    if theme == "Healthcare / Pharma" and any(term in evidence for term in ["clinical", "fda", "trial"]):
        return "Clinical catalyst; relevance depends on holding/watchlist"
    if "metadata match only" in evidence:
        return "Needs source PDF verification"
    if str(row.get("cutler_priority") or "") in {"Highest", "High"} or ticker in watchlist:
        return "Needs PM review"
    return "Monitor, no action"


def _cutler_priority_calibration(
    row: Dict[str, Any],
    theme: str,
    watchlist: set[str],
) -> Tuple[str, int, str]:
    """Rank a brief item by Cutler strategy relevance, independent of Dashboard signal priority."""
    text = " ".join(
        str(row.get(column) or "")
        for column in [
            "ticker", "company_or_identifier", "source_file", "file_name", "relative_path",
            "category", "document_type", "signal_type", "broker_or_source",
            "source_or_broker", "evidence",
        ]
    ).casefold()
    ticker = str(row.get("ticker") or "").upper()
    category = str(row.get("category") or "").casefold()
    document_type = str(row.get("document_type") or "").casefold()
    signal_direction = str(row.get("signal_direction") or "")
    watchlist_hit = bool(ticker and ticker in watchlist)
    weekly_or_guide = any(
        term in text
        for term in [
            "weekly", "monday", "jpm guide", "guide to the markets", "market recap",
            "convert vantage", "pricing report", "pricing toolkit", "reit toolkit",
        ]
    )
    community_bank = category == "banks" or any(
        term in text for term in ["community bank", "bank earnings", "bank investor", "bank shareholder"]
    )
    bank_primary_material = community_bank and (
        document_type in {
            "8-k/8k", "10-k/10k", "earnings release", "earnings presentation", "earnings supplement",
            "prepared remarks", "transcript", "current report",
        }
        or any(term in text for term in ["investor presentation", "shareholder", "8-k", "8k", "10-k", "10k"])
    )
    green_street = "green street" in text
    reit_real_estate = theme == "REITs / Real Estate"
    convertible_relevant = any(
        term in text
        for term in ["convertible", "convertibles", "convert vantage", "convert issuance", "convert offering"]
    )
    credit_concern = any(
        term in text
        for term in [
            "downgrade", "liquidity pressure", "liquidity concern", "leverage concern",
            "leverage pressure", "issuance need", "issuance pressure", "rating pressure",
        ]
    ) or (str(row.get("signal_type") or "") == "Credit / Liquidity" and signal_direction == "Negative")
    industry_report = "industry report" in document_type or "sector report" in document_type
    strategy_industry = industry_report and any(
        term in text
        for term in [
            "reit", "real estate", "community bank", "banks", "financials", "convertible",
            "convertibles", "dividend", "income",
        ]
    )

    if weekly_or_guide:
        return "Medium", 58 + (4 if watchlist_hit else 0), "Recurring weekly/guide material; useful context but not a first-pass must-have."
    if bank_primary_material:
        return "Highest", 100, "Community-bank primary material is directly relevant to Cutler strategy review."
    if community_bank:
        return "Highest", 97, "Community-bank research is a core Cutler strategy priority."
    if green_street:
        return "Highest", 96, "Green Street research is a core REIT/real-estate priority."
    if reit_real_estate and industry_report:
        return "Highest", 94, "REIT/real-estate industry research is directly relevant to Cutler strategy review."
    if reit_real_estate:
        return "High", 88 + (4 if watchlist_hit else 0), "Single-name REIT/real-estate research is relevant to holdings and NAV review."
    if convertible_relevant:
        return "High", 87 + (4 if watchlist_hit else 0), "Explicit convertible relevance connects the item to Cutler strategy."
    if strategy_industry:
        return "High", 84, "Industry research directly intersects with a Cutler strategy."
    if theme == "Convertibles / Credit":
        if watchlist_hit or credit_concern:
            return "High", 82 + (4 if watchlist_hit else 0), "Credit evidence has a watchlist hit or explicit downside-risk concern."
        return "Lower", 34, "Generic credit research is lower priority without convertible or direct strategy relevance."
    if industry_report:
        return "Medium", 62, "Broad industry research may provide context but lacks a direct Cutler strategy hit."
    if watchlist_hit:
        return "High", 80, "Watchlist/portfolio-linked item requiring source-document review."
    if theme == "Healthcare / Pharma":
        return "Lower", 32, "Healthcare/pharma is lower priority without a watchlist or portfolio hit."
    if theme == "Insider Trading / Screens":
        return "Lower", 28, "Screening material is lower priority without a watchlist or portfolio hit."
    if theme == "Macro / Market Backdrop":
        return "Medium", 56, "Macro context is useful for triage but usually not a first-pass must-have."
    return "Lower", 30, "Generic single-name research is lower priority without a direct Cutler strategy or watchlist hit."


_CUTLER_DEEP_EVIDENCE_PATTERNS = {
    "REIT / Real Estate": (
        r"\breits?\b", r"\bnav\b", r"\bcap rates?\b", r"\bnoi\b", r"\baffo\b",
        r"\bffo\b", r"\boccupancy\b", r"\brent growth\b", r"\bdividends?\b",
        r"\bgreen street\b", r"\bmalls?\b", r"\bresidential\b", r"\bsector update\b",
        r"\brecommendation\b", r"\bacquisition\b", r"\bdisposition\b",
        r"\bpreferred stock\b", r"\bspg\b", r"\bmac\b", r"\bskt\b", r"\bsimon\b",
        r"\bbay area\b", r"\bsun belt\b", r"\bmanufactured housing\b", r"\bvre\b",
        r"\bfcpt\b", r"\bvici\b", r"\bglpi\b", r"\bpeb\b", r"\baple\b", r"\byield\b",
    ),
    "Community Banks / Financials": (
        r"\bbanks?\b", r"\bcommunity banks?\b", r"\bdeposits?\b", r"\bnim\b",
        r"\bnet interest margin\b",
        r"\bcapital ratios?\b", r"\bloan growth\b", r"\basset quality\b", r"\bm&a\b",
        r"\bmergers?\b", r"\bdividends?\b",
    ),
    "Convertibles / Credit": (
        r"\bconvertible notes?\b", r"\bconvertibles?\b", r"\bcapped calls?\b",
        r"\bdilution\b", r"\bissuance\b", r"\buse of proceeds\b", r"\bleverage\b",
        r"\bliquidity\b", r"\bdebt\b", r"\bconverts? advanced\b", r"\bconverts? declined\b",
        r"\btreasury\b", r"\boil\b", r"\bequity rally\b", r"\bai infrastructure\b",
        r"\bspace\s*/?\s*defense\b", r"\b4\.125%\b", r"\$75m\b", r"\$300m\b",
    ),
    "Macro / Weekly": (
        r"\bweekly\b", r"\bjpm guide\b", r"\bguide to the markets\b", r"\bmarket recap\b",
        r"\bfear\s+(?:and|&)\s+greed\b", r"\bextreme greed\b", r"\bgreed\b",
        r"\bmajor indexes?\b", r"\bequities rallied\b", r"\bmarch bottom\b",
        r"\bai suppliers?\b", r"\bai customers?\b", r"\bpayrolls?\b", r"\bnonfarm payrolls?\b",
        r"\bpmi\b",
    ),
    "Healthcare / Clinical Catalyst": (
        r"\basco\b", r"\bclinical\b", r"\btrials?\b", r"\bkeytruda\b",
        r"\bmelanoma\b", r"\bbraftovi\b", r"\bmoderna\b", r"\bmrna\b",
        r"\bintismeran autogene\b", r"\brecurrence\b", r"\bdeath risk\b",
        r"\bmetastasis\b", r"\bsac-tmt\b", r"\boptitrop\b", r"\bprogression-free survival\b",
    ),
    "Insider Trading Screen": (
        r"\bopeninsider\b", r"\binsider trading\b", r"\binsider buying\b",
        r"\binsider selling\b", r"\bcrwd\b", r"\bamzn\b", r"\bfwonk\b", r"\bexp\b",
    ),
}

_CUTLER_TARGETED_REIT_TICKERS = {"FCPT", "VICI", "APLE", "PEB", "VRE", "GLPI", "SPG", "MAC", "SKT", "KIM"}
_CUTLER_TARGETED_HEALTHCARE_TICKERS = {"MRK", "PFE", "MRNA"}
_CUTLER_TARGETED_BANK_TICKERS = {"BORT", "FCCO"}


def _cutler_metadata_text(row: Dict[str, Any]) -> str:
    return " ".join(
        str(row.get(column) or "")
        for column in [
            "ticker", "company_or_identifier", "source_file", "file_name", "relative_path",
            "category", "document_type", "signal_type", "broker_or_source",
            "source_or_broker", "evidence",
        ]
    )


def cutler_daily_brief_document_type(row: Dict[str, Any]) -> str:
    """Normalize display document types for the Cutler daily packet only."""
    text = _normalize_metadata_text(_cutler_metadata_text(row)).casefold()
    category = str(row.get("category") or "").casefold()
    source = str(row.get("source_or_broker") or row.get("broker_or_source") or "").casefold()
    if re.search(r"\b10\s*k\b", text):
        return "Filing / 10-K"
    if re.search(r"\b8\s*k\b", text):
        return "Filing / 8-K"
    if re.search(r"\b(?:asco|clinical|trials?|keytruda|melanoma|braftovi)\b", text):
        return "Healthcare / Clinical Catalyst"
    if any(term in text for term in ["convertible note", "convertible notes", "senior notes", "capped call"]):
        return "Convertible Financing"
    if any(term in text for term in ["fear and greed", "fear & greed"]):
        return "Macro Sentiment"
    if any(term in text for term in ["jpm weekly", "jpm guide", "monday guide", "market recap"]):
        return "Macro / Weekly Guide"
    if re.search(r"\b(?:2q|q[1-4]|earnings)\s+(?:earnings\s+)?reminder\b", text):
        return "Company Event / Earnings Reminder"
    if "insider trading" in text:
        return "Insider Trading Screen"
    if any(term in text for term in ["investor ppt", "investor presentation", "presentation", " deck"]):
        return "Investor Presentation / REIT" if re.search(r"\b(?:aple|reit|real estate)\b", text) else "Investor Presentation"
    if re.search(r"\bhci\b", text) and any(term in text for term in ["catastrophe", "reinsurance"]):
        return "Company Press Release / Insurance Risk Update"
    if category == "green street" or source == "green street" or "green street" in text:
        return "Green Street / Sector Report"
    if any(
        term in text
        for term in ["mall sector", "residential", "reit pricing", "reit toolkit", "real estate report"]
    ):
        return "REIT / Real Estate Report"
    if "transcript" in text:
        return "Earnings Transcript"
    if re.search(r"\bsiri\b", text) and any(term in text for term in ["podcast", "comedy", "deal"]):
        return "Company News"
    if any(term in text for term in ["press release", "announces", "completion", "acquisition"]):
        return "Company Press Release"
    if "industry" in text:
        return "Industry Report"
    if category == "credit" or "credit" in text or source == "gimme credit":
        return "Credit Report"
    raw = str(row.get("document_type") or row.get("signal_type") or "").strip()
    if raw and raw.casefold() not in {"other", "unclassified document"}:
        return raw
    if any(term in text for term in ["reminder", "event"]):
        return "Company Event"
    if any(term in text for term in ["podcast", "deal", "update", "news"]):
        return "Company News"
    return "Company / Research Update"


def cutler_daily_brief_topic(row: Dict[str, Any]) -> str:
    """Create a concise topic/entity label for the Cutler daily packet."""
    text = _normalize_metadata_text(_cutler_metadata_text(row))
    folded = text.casefold()
    if "fear and greed" in folded or "fear & greed" in folded:
        return "Fear & Greed Index"
    if any(term in folded for term in ["jpm weekly", "jpm guide", "monday guide", "guide to the markets"]):
        return "JPM Weekly Market Guide"
    if "insider trading" in folded:
        return "Insider Trading Screen"
    if "green street" in folded:
        if "vre" in folded and any(term in folded for term in ["dropping research coverage", "coverage dropped"]):
            return "VRE / Coverage Dropped"
        if "qt" in folded and "fertitta" in folded and "caesars" in folded:
            return "QT / Caesars-Fertitta Read-Through"
        if "mall sector" in folded:
            return "Mall Sector Update"
        if "residential" in folded:
            return "Residential Sector Update"

    ticker = str(row.get("ticker") or "").strip().upper()
    if ticker and ticker.casefold() not in {"uncertain", "unclassified", "unknown", "n/a", "none"}:
        return ticker
    identifier = str(row.get("company_or_identifier") or "").strip()
    if identifier.casefold() in {"uncertain", "unclassified", "unknown", "n/a", "none"}:
        identifier = ""
    if identifier:
        cleaned = _normalize_metadata_text(identifier)
    else:
        file_name = str(row.get("source_file") or row.get("file_name") or "")
        cleaned = _normalize_metadata_text(Path(file_name).stem)
    cleaned = re.sub(
        r"\b(?:green street|jpmorgan|jpm|goldman sachs|wells fargo|gimme credit)\b",
        " ",
        cleaned,
        flags=re.IGNORECASE,
    )
    cleaned = re.sub(r"\b\d{1,2}\s+\d{1,2}\s+\d{2,4}\b|\b20\d{2}\b", " ", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" -_/")
    return cleaned[:100] if cleaned else "Unclassified"


def cutler_daily_brief_why_it_matters(row: Dict[str, Any], theme: str) -> str:
    """Return a concise strategy-specific rationale without making a financial conclusion."""
    document_type = str(row.get("cutler_document_type") or cutler_daily_brief_document_type(row))
    if document_type in {"Green Street / Sector Report", "REIT / Real Estate Report", "Investor Presentation / REIT"}:
        return "May affect REIT NAV assumptions, sector positioning, or PM review priorities."
    if theme == "Community Banks / Financials":
        return "Directly relevant to Cutler's community bank strategy review."
    if document_type == "Convertible Financing":
        return "May affect dilution, balance sheet, or convertible-security review."
    if document_type in {"Macro Sentiment", "Macro / Weekly Guide"} or theme == "Macro / Market Backdrop":
        return "Provides market backdrop for daily research triage."
    if document_type == "Industry Report":
        return "Provides sector-level read-through for covered/watchlist names."
    if document_type == "Healthcare / Clinical Catalyst":
        return "Clinical catalyst; relevance depends on holding/watchlist exposure."
    if document_type == "Insider Trading Screen":
        return "Screening item only; do not treat as a conclusion without watchlist/holding link."
    if document_type == "Company Event / Earnings Reminder":
        return "Flags an upcoming earnings event for watchlist/PM awareness."
    if document_type in {
        "Company News", "Company Press Release", "Company Press Release / Insurance Risk Update",
        "Company / Research Update", "Company Event",
    }:
        return "Company-specific update; review only if portfolio/watchlist relevant."
    if document_type.startswith("Filing /"):
        return "Primary-source filing for portfolio/watchlist or strategy review."
    if document_type.startswith("Investor Presentation"):
        return "Primary company presentation for portfolio/watchlist or strategy review."
    if theme == "REITs / Real Estate":
        return "May affect REIT NAV assumptions, sector positioning, or PM review priorities."
    return "Company-specific research item; review only if portfolio/watchlist relevant."


def best_cutler_daily_brief_evidence(item: Dict[str, Any], *, max_chars: int = 900) -> Dict[str, Any]:
    """Return a useful Cutler-theme passage from limited or deeper extracted text."""
    status = str(item.get("text_extraction_status") or "")
    text = _normalized_extracted_text(item)
    result = {"snippet": "", "evidence_type": "Other", "quality": "unavailable"}
    if status != "scanned" or not text:
        return result

    sentences = [
        sentence.strip()
        for sentence in re.split(r"(?<=[.!?])\s+|\s{2,}", text)
        if sentence.strip()
    ] or [text]
    candidates = []
    for index in range(len(sentences)):
        for window_size in (1, 2, 3):
            passage = " ".join(sentences[index:index + window_size]).strip()[:max_chars]
            if not passage:
                continue
            scored_types = []
            for evidence_type, patterns in _CUTLER_DEEP_EVIDENCE_PATTERNS.items():
                hits = sum(bool(re.search(pattern, passage, flags=re.IGNORECASE)) for pattern in patterns)
                if hits:
                    scored_types.append((hits, evidence_type))
            useful_hits = sum(hits for hits, _ in scored_types)
            disclosure_hits = sum(phrase in passage.casefold() for phrase in DISCLOSURE_PHRASES)
            score = (useful_hits * 8) - (disclosure_hits * 18) + min(len(passage), max_chars) / max_chars
            evidence_type = max(scored_types, default=(0, "Other"))[1]
            candidates.append((score, useful_hits, passage, evidence_type))
    best = max(candidates, default=None, key=lambda value: value[0])
    if best and best[1] > 0 and best[0] > 0:
        return {
            "snippet": best[2],
            "evidence_type": best[3],
            "quality": "cutler_useful",
        }
    generic = best_investment_useful_snippet(item, max_chars=max_chars)
    return {
        "snippet": str(generic.get("snippet") or ""),
        "evidence_type": str(generic.get("evidence_type") or "Other"),
        "quality": str(generic.get("quality") or "no_useful_snippet"),
    }


def extract_cutler_daily_brief_facts(item: Dict[str, Any], *, max_facts: int = 4) -> List[str]:
    """Select concise, source-verbatim fact sentences from targeted extracted text."""
    if str(item.get("text_extraction_status") or "") != "scanned":
        return []
    text = _normalized_extracted_text(item)
    if not text:
        return []
    candidates = []
    for sentence in re.split(r"(?<=[.!?])\s+|\s{2,}", text):
        cleaned = re.sub(r"\s+", " ", sentence).strip(" -•\t")
        if len(cleaned) < 25 or any(phrase in cleaned.casefold() for phrase in DISCLOSURE_PHRASES):
            continue
        matched_types = []
        hit_count = 0
        for evidence_type, patterns in _CUTLER_DEEP_EVIDENCE_PATTERNS.items():
            hits = sum(bool(re.search(pattern, cleaned, flags=re.IGNORECASE)) for pattern in patterns)
            if hits:
                matched_types.append(evidence_type)
                hit_count += hits
        if not hit_count:
            continue
        number_boost = 2 if re.search(r"(?:\$[\d,.]+|\b\d+(?:\.\d+)?%|\b\d+(?:\.\d+)?x\b)", cleaned) else 0
        candidates.append((hit_count + number_boost, matched_types, cleaned[:300]))
    selected = []
    seen = set()
    represented_types = set()
    for _, types, sentence in sorted(candidates, key=lambda value: (-value[0], len(value[2]))):
        signature = re.sub(r"\W+", " ", sentence.casefold()).strip()
        if signature in seen:
            continue
        if selected and set(types) <= represented_types and len(selected) >= max_facts - 1:
            continue
        selected.append(sentence)
        seen.add(signature)
        represented_types.update(types)
        if len(selected) >= max_facts:
            break
    return selected


def _is_targeted_cutler_deep_file(row: Dict[str, Any]) -> bool:
    text = _normalize_metadata_text(_cutler_metadata_text(row)).casefold()
    ticker = str(row.get("ticker") or "").upper()
    category = str(row.get("category") or "").casefold()
    return (
        category == "green street"
        or ticker in _CUTLER_TARGETED_REIT_TICKERS
        or ticker in _CUTLER_TARGETED_HEALTHCARE_TICKERS
        or ticker in _CUTLER_TARGETED_BANK_TICKERS
        or any(
            term in text
            for term in [
                "jpm reit", "reit toolkit", "reit pricing", "convert vantage", "odv convertible",
                "fear and greed", "jpm weekly", "jpm guide", "community bank", "openinsider",
                "insider trading",
            ]
        )
    )


def select_cutler_daily_brief_files(
    relevance: pd.DataFrame,
    *,
    max_files: int,
    watchlist: Iterable[str] = (),
) -> pd.DataFrame:
    """Select only the highest-value files for the Daily Brief deeper extraction pass."""
    if not isinstance(relevance, pd.DataFrame) or relevance.empty:
        return relevance.iloc[0:0].copy() if isinstance(relevance, pd.DataFrame) else pd.DataFrame()
    watchlist_set = {str(value).strip().upper() for value in watchlist if str(value).strip()}
    eligible = relevance[
        relevance["file_extension"].astype(str).isin(SUPPORTED_TEXT_EXTENSIONS)
        & relevance["extraction_status"].astype(str).eq("extracted")
    ].copy()
    if eligible.empty:
        return eligible

    ranked = []
    priority_rank = {"Highest": 0, "High": 1, "Medium": 2, "Lower": 3}
    for _, row in eligible.iterrows():
        record = row.to_dict()
        theme = _cutler_brief_theme(record)
        priority, score, reason = _cutler_priority_calibration(record, theme, watchlist_set)
        text = " ".join(
            str(record.get(column) or "")
            for column in ["ticker", "file_name", "relative_path", "category", "document_type"]
        ).casefold()
        explicit_odv_convertible = "odv" in text and any(
            term in text for term in ["convertible", "convertible note", "convert note"]
        )
        targeted_file = _is_targeted_cutler_deep_file(record)
        strategy_candidate = (
            priority in {"Highest", "High"}
            or targeted_file
            or explicit_odv_convertible
            or any(
                term in text
                for term in [
                    "green street", "reit", "real estate", "convertible", "convert note",
                    "community bank", "industry report", "industry update", "pricing toolkit",
                    "pricing report",
                ]
            )
            or (str(record.get("ticker") or "").upper() in watchlist_set)
        )
        if not strategy_candidate:
            continue
        record["_cutler_priority_rank"] = min(priority_rank.get(priority, 4), 1 if targeted_file else 4)
        record["_cutler_deep_score"] = score + (18 if targeted_file else 0) + (8 if explicit_odv_convertible else 0)
        record["cutler_priority"] = priority
        record["cutler_priority_reason"] = reason
        record["targeted_deep_extraction"] = targeted_file
        record["cutler_topic"] = cutler_daily_brief_topic(record)
        record["cutler_document_type"] = cutler_daily_brief_document_type(record)
        ranked.append(record)
    if not ranked:
        return eligible.iloc[0:0].copy()
    return (
        pd.DataFrame(ranked)
        .sort_values(
            ["_cutler_priority_rank", "_cutler_deep_score", "relevance_score", "file_name"],
            ascending=[True, False, False, True],
        )
        .drop_duplicates(subset=["relative_path"])
        .head(int(max_files))
        .drop(columns=["_cutler_priority_rank", "_cutler_deep_score"])
        .reset_index(drop=True)
    )


def prepare_cutler_daily_brief_text(
    selected: pd.DataFrame,
    existing_text: List[Dict[str, Any]],
    *,
    max_pdf_pages: int,
    max_chars_per_file: int,
) -> List[Dict[str, Any]]:
    """Deeply rescan selected Cutler-priority files and merge them over light snippets."""
    existing_by_path = {
        str(item.get("relative_path") or ""): dict(item)
        for item in existing_text or []
        if str(item.get("relative_path") or "")
    }
    if not isinstance(selected, pd.DataFrame) or selected.empty:
        return list(existing_by_path.values())
    deeper = extract_selected_text(
        selected,
        max_pdf_pages=max_pdf_pages,
        max_chars_per_file=max_chars_per_file,
    )
    for item in deeper:
        path = str(item.get("relative_path") or "")
        prior = existing_by_path.get(path)
        if str(item.get("text_extraction_status") or "") != "scanned" and prior:
            prior["deeper_extraction_attempt_status"] = item.get("text_extraction_status") or "not_scanned"
            continue
        item["extraction_depth"] = "deeper"
        evidence = best_cutler_daily_brief_evidence(item, max_chars=1200)
        item["cutler_evidence_snippet"] = evidence["snippet"]
        item["cutler_evidence_type"] = evidence["evidence_type"]
        item["cutler_evidence_quality"] = evidence["quality"]
        item["cutler_extracted_facts"] = extract_cutler_daily_brief_facts(item, max_facts=5)
        existing_by_path[path] = item
    for item in existing_by_path.values():
        item.setdefault("extraction_depth", "limited")
        if (
            str(item.get("text_extraction_status") or "") == "scanned"
            and not str(item.get("cutler_evidence_snippet") or "").strip()
        ):
            evidence = best_cutler_daily_brief_evidence(item, max_chars=900)
            item["cutler_evidence_snippet"] = evidence["snippet"]
            item["cutler_evidence_type"] = evidence["evidence_type"]
            item["cutler_evidence_quality"] = evidence["quality"]
        if not item.get("cutler_extracted_facts"):
            item["cutler_extracted_facts"] = extract_cutler_daily_brief_facts(item, max_facts=4)
    return list(existing_by_path.values())


def build_cutler_style_daily_brief(
    relevance: pd.DataFrame,
    selected_text: List[Dict[str, Any]],
    *,
    source_name: str,
    signals: Optional[pd.DataFrame] = None,
    broker_coverage: Optional[pd.DataFrame] = None,
    manual_review: Optional[pd.DataFrame] = None,
    watchlist: Iterable[str] = (),
    mode: str = "Fast",
) -> str:
    """Build a deterministic, source-cited daily packet organized around Cutler themes."""
    if not isinstance(relevance, pd.DataFrame):
        relevance = pd.DataFrame(columns=RELEVANCE_COLUMNS)
    if not isinstance(signals, pd.DataFrame):
        index_rows = build_research_index_rows(relevance, selected_text, source_name=source_name)
        signals = build_dashboard_signal_table(index_rows)
    if not isinstance(broker_coverage, pd.DataFrame):
        broker_coverage = build_broker_coverage_summary(relevance)
    if not isinstance(manual_review, pd.DataFrame):
        manual_review = build_dashboard_manual_review_queue(signals)

    source_meta = {}
    for _, row in relevance.iterrows():
        file_name = str(row.get("file_name") or Path(str(row.get("relative_path") or "")).name)
        if file_name:
            source_meta[file_name] = row.to_dict()
    selected_by_file = {}
    for item in selected_text or []:
        file_name = str(item.get("file_name") or Path(str(item.get("relative_path") or "")).name)
        if file_name:
            selected_by_file[file_name] = {
                "evidence": str(item.get("cutler_evidence_snippet") or "").strip(),
                "evidence_type": str(item.get("cutler_evidence_type") or ""),
                "extraction_depth": str(item.get("extraction_depth") or "limited"),
                "text_extraction_status": str(item.get("text_extraction_status") or ""),
                "facts": list(item.get("cutler_extracted_facts") or []),
            }
    cutler_priority_rank = {"Highest": 0, "High": 1, "Medium": 2, "Lower": 3}
    max_per_section = 10 if str(mode).casefold().startswith("deeper") else 5
    max_must_have = 10 if str(mode).casefold().startswith("deeper") else 6
    watchlist_set = {str(value).strip().upper() for value in watchlist if str(value).strip()}
    records = []
    for _, signal in signals.iterrows():
        record = signal.to_dict()
        source_file = str(record.get("source_file") or "")
        metadata = source_meta.get(source_file, {})
        for column in ["category", "document_type", "company_or_identifier", "ticker"]:
            if not str(record.get(column) or "").strip():
                record[column] = metadata.get(column) or ""
        record["relevance_score"] = int(metadata.get("relevance_score") or 0)
        extracted = selected_by_file.get(source_file, {})
        if extracted.get("evidence"):
            record["evidence"] = extracted["evidence"]
            record["signal_type"] = extracted.get("evidence_type") or record.get("signal_type") or "Other"
            record["confidence"] = "High"
        record["extraction_depth"] = extracted.get("extraction_depth") or "limited"
        record["cutler_extracted_facts"] = list(extracted.get("facts") or [])
        record["theme"] = _cutler_brief_theme(record)
        record["cutler_topic"] = cutler_daily_brief_topic(record)
        record["cutler_document_type"] = cutler_daily_brief_document_type(record)
        record["cutler_why_it_matters"] = cutler_daily_brief_why_it_matters(record, record["theme"])
        (
            record["cutler_priority"],
            record["cutler_relevance_score"],
            record["cutler_priority_reason"],
        ) = _cutler_priority_calibration(record, record["theme"], watchlist_set)
        record["actionability"] = _cutler_actionability(record, record["theme"], watchlist_set)
        records.append(record)

    represented_sources = {str(record.get("source_file") or "") for record in records}
    for _, row in relevance.sort_values("relevance_score", ascending=False).iterrows():
        file_name = str(row.get("file_name") or Path(str(row.get("relative_path") or "")).name)
        if not file_name or file_name in represented_sources:
            continue
        extracted = selected_by_file.get(file_name, {})
        extracted_evidence = str(extracted.get("evidence") or "")
        record = {
            "ticker": row.get("ticker") or "",
            "company_or_identifier": row.get("company_or_identifier") or "",
            "signal_type": extracted.get("evidence_type") or row.get("document_type") or "Other",
            "signal_direction": "Unknown",
            "confidence": "High" if extracted_evidence else "Low",
            "priority": "Medium" if str(row.get("priority_level") or "") == "High" else "Low",
            "evidence": extracted_evidence or (
                f"Metadata match only: {row.get('document_type') or 'other'} in "
                f"{row.get('category') or 'Root'}."
            ),
            "source_file": file_name,
            "broker_or_source": row.get("source_or_broker") or "",
            "category": row.get("category") or "",
            "document_type": row.get("document_type") or "",
            "why_it_matters": "Helps prioritize source-document review.",
            "why_review": "Metadata-only item; review the source PDF for details.",
            "relevance_score": int(row.get("relevance_score") or 0),
            "extraction_depth": extracted.get("extraction_depth") or "limited",
            "cutler_extracted_facts": list(extracted.get("facts") or []),
        }
        record["theme"] = _cutler_brief_theme(record)
        record["cutler_topic"] = cutler_daily_brief_topic(record)
        record["cutler_document_type"] = cutler_daily_brief_document_type(record)
        record["cutler_why_it_matters"] = cutler_daily_brief_why_it_matters(record, record["theme"])
        (
            record["cutler_priority"],
            record["cutler_relevance_score"],
            record["cutler_priority_reason"],
        ) = _cutler_priority_calibration(record, record["theme"], watchlist_set)
        record["actionability"] = _cutler_actionability(record, record["theme"], watchlist_set)
        records.append(record)

    def evidence_rank(item: Dict[str, Any]) -> int:
        evidence = str(item.get("evidence") or "").casefold()
        if str(item.get("extraction_depth") or "") == "deeper" and "metadata match only" not in evidence:
            return 0
        if "metadata match only" not in evidence:
            return 1
        return 2

    records.sort(
        key=lambda item: (
            cutler_priority_rank.get(str(item.get("cutler_priority") or ""), 4),
            evidence_rank(item),
            -int(item.get("cutler_relevance_score") or 0),
            str(item.get("source_file") or ""),
        )
    )
    thematic_sections = CUTLER_BRIEF_SECTIONS[1:9]
    by_theme = {section: [] for section in thematic_sections}
    seen_theme_sources = set()
    for record in records:
        theme = str(record.get("theme") or "Other Notable Research")
        key = (theme, str(record.get("source_file") or ""))
        if theme not in by_theme or key in seen_theme_sources or len(by_theme[theme]) >= max_per_section:
            continue
        by_theme[theme].append(record)
        seen_theme_sources.add(key)

    must_have_candidates = [
        record for record in records
        if str(record.get("cutler_priority") or "") in {"Highest", "High"}
    ]
    must_have = []
    seen_must_have_sources = set()
    for record in must_have_candidates:
        source = str(record.get("source_file") or "")
        if not source or source in seen_must_have_sources:
            continue
        must_have.append(record)
        seen_must_have_sources.add(source)
        if len(must_have) >= max_must_have:
            break
    if not must_have:
        for record in records:
            source = str(record.get("source_file") or "")
            if not source or source in seen_must_have_sources:
                continue
            must_have.append(record)
            seen_must_have_sources.add(source)
            if len(must_have) >= max_must_have:
                break

    def source_citation(sources: Iterable[str]) -> str:
        unique = list(dict.fromkeys(str(source) for source in sources if str(source).strip()))
        if len(unique) == 1:
            return f"[Source: {unique[0]}]"
        return f"[Sources: {'; '.join(unique)}]"

    def evidence_summary(record: Dict[str, Any]) -> str:
        evidence = re.sub(r"\s+", " ", str(record.get("evidence") or "")).strip()
        document_type = str(record.get("cutler_document_type") or "Company / Research Update")
        if "metadata match only" in evidence.casefold() or not evidence:
            return f"Metadata-only: a {document_type} file is present and relevant for review."
        facts = [
            re.sub(r"\s+", " ", str(fact)).strip().rstrip(".")
            for fact in record.get("cutler_extracted_facts") or []
            if str(fact).strip()
        ]
        if facts:
            label = "Targeted deeper extraction" if str(record.get("extraction_depth") or "") == "deeper" else "Limited extraction"
            return f"{label} identifies " + "; ".join(facts[:3]) + "."
        sentences = [
            sentence.strip()
            for sentence in re.split(r"(?<=[.!?])\s+", evidence)
            if sentence.strip()
        ]
        unique_sentences = list(dict.fromkeys(sentences))
        excerpt = " ".join(unique_sentences[:2])[:360].strip()
        if len(" ".join(unique_sentences[:2])) > 360:
            excerpt += "..."
        label = "Targeted deeper extraction identifies" if str(record.get("extraction_depth") or "") == "deeper" else "Limited extraction identifies"
        return f"{label} {excerpt.rstrip('.')}."

    def section_bullet(record: Dict[str, Any]) -> str:
        file_name = str(record.get("source_file") or "source unavailable")
        topic = str(record.get("cutler_topic") or "Unclassified")
        document_type = str(record.get("cutler_document_type") or "Company / Research Update")
        priority = str(record.get("cutler_priority") or "Lower")
        rationale = str(record.get("cutler_why_it_matters") or "Supports source-document triage.")
        action = str(record.get("actionability") or "Monitor, no action")
        return (
            f"- **{topic}:** {evidence_summary(record)} {rationale} "
            f"**{priority} priority.** Actionability: {action}. {source_citation([file_name])}"
        )

    def executive_group_key(record: Dict[str, Any]) -> Tuple[str, str]:
        theme = str(record.get("theme") or "")
        topic = str(record.get("cutler_topic") or "Unclassified")
        ticker = str(record.get("ticker") or "").upper()
        if theme == "Community Banks / Financials":
            return ("theme", "Community Banks")
        if theme == "REITs / Real Estate":
            return ("reit", topic.casefold())
        if ticker and ticker not in {"UNCERTAIN", "UNCLASSIFIED", "UNKNOWN"}:
            return ("ticker", ticker)
        return ("topic", topic.casefold())

    executive_groups: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
    for record in must_have:
        executive_groups.setdefault(executive_group_key(record), []).append(record)

    def executive_bullet(group: List[Dict[str, Any]]) -> str:
        ordered = sorted(
            group,
            key=lambda item: (
                evidence_rank(item),
                cutler_priority_rank.get(str(item.get("cutler_priority") or ""), 4),
                -int(item.get("cutler_relevance_score") or 0),
            ),
        )
        lead = ordered[0]
        sources = [str(item.get("source_file") or "") for item in ordered]
        if str(lead.get("theme") or "") == "Community Banks / Financials":
            topic_docs = []
            seen_topics = set()
            for item in ordered:
                topic = str(item.get("cutler_topic") or "Unclassified")
                if topic in seen_topics:
                    continue
                topic_docs.append(f"{topic} has a {item.get('cutler_document_type') or 'research file'}")
                seen_topics.add(topic)
            coverage = "; ".join(topic_docs[:4])
            evidence_note = ""
            if evidence_rank(lead) < 2:
                evidence_note = " " + evidence_summary(lead)
            return (
                f"- **Community Banks:** {coverage}. Directly relevant to Cutler's community bank strategy review."
                f"{evidence_note} {source_citation(sources)}"
            )
        topic = str(lead.get("cutler_topic") or "Unclassified")
        rationale = str(lead.get("cutler_why_it_matters") or "Supports source-document triage.")
        return f"- **{topic}:** {evidence_summary(lead)} {rationale} {source_citation(sources)}"

    source_suffix = daily_source_title_suffix(source_name, relevance.get("relative_path", []))
    lines = [
        f"# Cutler-Style Daily Research Brief - {source_suffix}",
        "",
        f"Source archive: `{source_name}`",
        (
            "This packet is based on inventory metadata and limited extracted snippets. "
            "It does not claim full-PDF review."
        ),
        "",
        "## Executive Summary",
    ]
    if must_have:
        for group in list(executive_groups.values())[:5]:
            lines.append(executive_bullet(group))
    else:
        lines.append("No source-grounded research items were available for this packet.")

    for section in thematic_sections:
        lines.extend(["", f"## {section}"])
        if by_theme[section]:
            lines.extend(section_bullet(record) for record in by_theme[section])
        else:
            lines.append("No qualifying source-grounded items were identified for this section.")

    definite_groups = [
        ("Top REIT Takeaways", lambda item: str(item.get("theme") or "") == "REITs / Real Estate"),
        (
            "Convertible Market and ODV",
            lambda item: str(item.get("theme") or "") == "Convertibles / Credit"
            and (
                "convert" in _cutler_metadata_text(item).casefold()
                or str(item.get("ticker") or "").upper() == "ODV"
            ),
        ),
        ("Healthcare Catalysts", lambda item: str(item.get("theme") or "") == "Healthcare / Pharma"),
        ("Macro Frame", lambda item: str(item.get("theme") or "") == "Macro / Market Backdrop"),
        ("Insider Trading Screen", lambda item: str(item.get("theme") or "") == "Insider Trading / Screens"),
    ]
    lines.extend(["", "## What the Report Should Definitely Catch"])
    included_catch_sources = set()
    for heading, matcher in definite_groups:
        group_records = [
            record for record in records
            if matcher(record) and str(record.get("source_file") or "") not in included_catch_sources
        ][:4]
        if not group_records:
            continue
        lines.extend(["", f"### {heading}"])
        for record in group_records:
            lines.append(section_bullet(record))
            included_catch_sources.add(str(record.get("source_file") or ""))
    lines.extend(["", "### Actionability"])
    if must_have:
        for record in must_have[:6]:
            lines.append(
                f"- {record.get('actionability') or 'Monitor, no action'}: "
                f"{record.get('cutler_topic') or 'Unclassified'} requires source-document verification. "
                f"{source_citation([record.get('source_file') or 'source unavailable'])}"
            )
    else:
        lines.append("No source-grounded actionability items were identified.")

    lines.extend(["", "## Must-Have Items for Review"])
    if must_have:
        for record in must_have:
            lines.append(
                f"- **{record.get('cutler_topic') or 'Unclassified'} — "
                f"{record.get('cutler_document_type') or 'Company / Research Update'}:** "
                f"{record.get('cutler_why_it_matters') or 'Supports source-document triage.'} "
                f"Suggested action: {record.get('actionability') or 'Needs PM review'}. "
                f"{source_citation([record.get('source_file') or 'source unavailable'])}"
            )
    else:
        lines.append("No must-have items were identified from the available evidence.")

    lines.extend(["", "## Actionability / Suggested Follow-Up"])
    if must_have:
        for record in must_have[:5]:
            lines.append(
                f"- {record.get('actionability') or 'Monitor, no action'}: verify "
                f"{record.get('cutler_topic') or 'the unclassified topic'} "
                f"{str(record.get('cutler_document_type') or 'research').casefold()} evidence in the source document. "
                f"[Source: {record.get('source_file') or 'source unavailable'}]"
            )
    else:
        lines.append("No source-specific follow-up was generated.")

    used_sources = []
    for section_records in by_theme.values():
        used_sources.extend(str(record.get("source_file") or "") for record in section_records)
    used_sources.extend(str(record.get("source_file") or "") for record in must_have)
    used_sources.extend(included_catch_sources)
    used_sources = list(dict.fromkeys(source for source in used_sources if source))
    lines.extend(["", "## Source References"])
    if used_sources:
        for source in used_sources:
            metadata = source_meta.get(source, {})
            extracted = selected_by_file.get(source, {})
            status = "metadata-only"
            if extracted.get("text_extraction_status") == "scanned":
                status = (
                    "deeper extracted evidence"
                    if extracted.get("extraction_depth") == "deeper" and extracted.get("evidence")
                    else "limited snippet"
                )
            lines.append(f"- [Source: {source}] — {status}; {metadata.get('category') or 'category uncertain'}")
    else:
        lines.append("No source references were available.")

    lines.extend(
        [
            "",
            "## Caveats / Manual Verification",
            (
                "Financial conclusions, ratings, targets, estimates, and directional interpretations must be "
                "verified directly in the cited source files."
            ),
            (
                "Metadata-only observations establish file availability and classification only. "
                "This packet does not provide an investment recommendation."
            ),
        ]
    )
    return "\n".join(lines)


def _normalized_extracted_text(item: Dict[str, Any]) -> str:
    return re.sub(r"\s+", " ", str(item.get("extracted_text") or "")).strip()


def verified_extracted_text_items(selected_text: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    verified = []
    for item in selected_text:
        status = str(item.get("text_extraction_status") or "")
        text = _normalized_extracted_text(item)
        if status != "scanned" or len(text) < 160:
            continue
        verified.append(item)
    return verified


def validate_llm_brief_grounding(text: str, selected_text: List[Dict[str, Any]]) -> bool:
    """Require sensitive claims to be quoted or cautious evidence-organizer observations."""
    verified = verified_extracted_text_items(selected_text)
    if not verified:
        return not any(re.search(pattern, text, flags=re.IGNORECASE) for pattern in SENSITIVE_FINANCE_PATTERNS)

    sensitive_patterns = [
        *SENSITIVE_FINANCE_PATTERNS,
        r"\bratings?\b", r"\bprice targets?\b", r"\bestimates?\b",
        r"\bfree cash flow\b", r"\bcash flow\b", r"\bleverage\b", r"\bdebt\b",
        r"\bcredit\b", r"\bupgrades?\b", r"\bdowngrades?\b",
    ]
    organizer_pattern = re.compile(
        r"\b(?:the\s+)?(?:extracted\s+)?snippets?\s+(?:from\s+.+?\s+)?"
        r"(?:mention|mentions|include|includes|contain|contains)\b",
        flags=re.IGNORECASE,
    )
    prohibited_interpretation_pattern = re.compile(
        r"\b(?:this suggests|reports? indicate|brokers? (?:noted|believe|conclude)|"
        r"investment case|performance was|appears (?:strong|weak)|is (?:bullish|bearish))\b",
        flags=re.IGNORECASE,
    )
    for line in text.splitlines():
        stripped = line.strip()
        if (
            not stripped
            or stripped.startswith("#")
            or bool(re.fullmatch(r"[\s|:-]+", stripped))
        ):
            continue
        matched_patterns = [
            pattern for pattern in sensitive_patterns
            if re.search(pattern, stripped, flags=re.IGNORECASE)
        ]
        if not matched_patterns:
            continue
        cited = [
            item for item in verified
            if str(item.get("relative_path") or "") in stripped
            or str(item.get("file_name") or "") in stripped
        ]
        if not cited:
            return False
        quoted_parts = [part.strip() for part in re.findall(r'"([^"]+)"', stripped) if part.strip()]
        quote_supported = any(
            any(
                len(quote) >= 8
                and quote.lower() in _normalized_extracted_text(item).lower()
                for quote in quoted_parts
            )
            for item in cited
        )
        if quote_supported:
            continue
        if prohibited_interpretation_pattern.search(stripped) or not organizer_pattern.search(stripped):
            return False
        cited_text = " ".join(_normalized_extracted_text(item) for item in cited)
        if not all(re.search(pattern, cited_text, flags=re.IGNORECASE) for pattern in matched_patterns):
            return False
    return True


def build_llm_evidence_payload(
    relevance: pd.DataFrame,
    selected_text: List[Dict[str, Any]],
    *,
    max_total_chars: int = 22000,
) -> str:
    records = []
    used = 0
    for item in selected_text:
        snippet = _normalized_extracted_text(item)[:1200]
        status = str(item.get("text_extraction_status") or "")
        verified = status == "scanned" and len(snippet) >= 160
        record = {
            "relative_path": item.get("relative_path"),
            "ticker": item.get("ticker"),
            "category": item.get("category"),
            "source_or_broker": item.get("source_or_broker"),
            "document_type": item.get("document_type"),
            "relevance_score": item.get("relevance_score"),
            "text_extraction_status": status,
            "verified_text_available": verified,
            "limited_text": snippet if verified else INSUFFICIENT_TEXT_NOTICE,
        }
        encoded = json.dumps(record, ensure_ascii=False)
        if used + len(encoded) > max_total_chars:
            break
        records.append(record)
        used += len(encoded)
    metadata = relevance[
        [
            "relative_path", "ticker", "all_detected_tickers", "category", "source_or_broker",
            "document_type", "relevance_score", "priority_level", "investment_rationale",
        ]
    ].head(30).to_dict(orient="records") if not relevance.empty else []
    broker_coverage = build_broker_coverage_summary(relevance).head(20).to_dict(orient="records")
    ticker_summary = build_ticker_summary(relevance).head(20).to_dict(orient="records")
    return json.dumps(
        {
            "grounding_rule": (
                "Metadata may establish file availability, ticker/source/category coverage, and document type only. "
                "Financial claims require verified_text_available=true and must cite that source filename."
            ),
            "selected_sources": records,
            "top_metadata": metadata,
            "broker_coverage": broker_coverage,
            "ticker_summary": ticker_summary,
        },
        ensure_ascii=False,
    )
